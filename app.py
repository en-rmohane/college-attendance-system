import csv
import os
from sqlalchemy import or_
import requests
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import calendar
import random
import smtplib
from sqlalchemy import func
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from sqlalchemy import text
from werkzeug.utils import secure_filename, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash

from fix_database import get_timetable_from_db, generate_monthly_attendance_excel, basedir, send_otp_email
from models import db, User, Student, Subject, ProfessorSubject, Attendance, AttendanceReport, PasswordResetOTP, \
    EmailLog, RGPVScheme, TimetableSlot, CurrentSemester, MidTermMarks, Notes, Notice, Test, Question, \
    TestAttempt, StudentAnswer, QuestionSection, Faculty

from datetime import datetime

def get_now():
    """Return current naive datetime (same as other code)"""
    return datetime.now()
import os
from datetime import datetime, timezone, timedelta
import pytz  # pip install pytz

# IST Timezone setup
IST = pytz.timezone('Asia/Kolkata')

def get_ist_time():
    """Get current IST time"""
    return datetime.now(IST)

def convert_to_ist(utc_dt):
    """Convert UTC datetime to IST"""
    if utc_dt.tzinfo is None:
        utc_dt = pytz.utc.localize(utc_dt)
    return utc_dt.astimezone(IST)

def convert_to_utc(ist_dt):
    """Convert IST datetime to UTC"""
    if ist_dt.tzinfo is None:
        ist_dt = IST.localize(ist_dt)
    return ist_dt.astimezone(pytz.utc)

# Initialize Flask app
IS_VERCEL = os.environ.get('VERCEL') == '1'
if IS_VERCEL:
    app = Flask(__name__, instance_path='/tmp')
else:
    app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 280,
    "pool_size": 5,
    "max_overflow": 2,
}


''''# Email Configuration
app.config['MAIL_SERVER'] = 'smtp-relay.brevo.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'sbitmstudy@gmail.com'
app.config['MAIL_PASSWORD'] = 'vhr1SkHnaPgq5Jjb'
app.config['MAIL_DEFAULT_SENDER'] = 'sbitmstudy@gmail.com'''
import os

def setup_database():
    db_url = os.environ.get('DATABASE_URL')

    if db_url:
        # Render ka URL mostly 'postgres://...'
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql+psycopg://", 1)
        elif db_url.startswith("postgresql://"):
            db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)


        if "sslmode=" not in db_url:
            if "?" in db_url:
                db_url += "&sslmode=require"
            else:
                db_url += "?sslmode=require"

        print(f"✅ Database configured: {db_url[:60]}...")
        return db_url

    # Local development - SQLite
    if IS_VERCEL:
        print("⚠️ WARNING: DATABASE_URL is not set on Vercel. App will use local SQLite which is READ-ONLY and will NOT save data.")
    
    db_url = "sqlite:///college_attendance.db"
    print(f"✅ Database configured (local): {db_url}")
    return db_url

app.config["SQLALCHEMY_DATABASE_URI"] = setup_database()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# ========== SAFE EXTENSION INITIALIZATION ==========
# Initialize extensions ONLY if not already initialized
if 'db' not in globals():
    db = SQLAlchemy(app)
else:
    # If db already exists, just init the app with it
    db.init_app(app)

if 'login_manager' not in globals():
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message_category = 'info'

if 'migrate' not in globals():
    migrate = Migrate(app, db)
@login_manager.user_loader
def load_user(user_id):
    """Required by Flask-Login"""
    try:
        return User.query.get(int(user_id))
    except Exception as e:
        print(f"Error loading user: {e}")
        return None


# ========== END EXTENSION INITIALIZATION ==========

def get_today_attendance_summary():
    """Get today's attendance summary for all subjects"""
    try:
        today = date.today()
        subject_ids = db.session.query(Attendance.subject_id).filter(
            Attendance.date == today
        ).distinct().all()

        stats = []
        for (subject_id,) in subject_ids:
            subject = Subject.query.get(subject_id)
            if not subject:
                continue

            allotment = ProfessorSubject.query.filter_by(subject_id=subject_id).first()
            professor = User.query.get(allotment.professor_id) if allotment else None

            present_count = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today,
                status='present'
            ).count()

            absent_count = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today,
                status='absent'
            ).count()

            total_students = present_count + absent_count
            percentage = round((present_count / total_students) * 100, 2) if total_students > 0 else 0

            latest_record = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today
            ).order_by(Attendance.created_at.desc()).first()

            last_updated = latest_record.created_at.strftime('%H:%M') if latest_record else 'N/A'

            stats.append({
                'subject_id': subject_id,
                'subject_code': subject.code,
                'subject_name': subject.name,
                'professor_name': professor.fullname if professor else 'Not Allotted',
                'branch': subject.branch,
                'year': (subject.semester + 1) // 2,
                'semester': subject.semester,
                'present_count': present_count,
                'absent_count': absent_count,
                'total_students': total_students,
                'percentage': percentage,
                'last_updated': last_updated
            })

        stats.sort(key=lambda x: x['last_updated'], reverse=True)
        return stats

    except Exception as e:
        print(f"Error in get_today_attendance_summary: {e}")
        return []
def preload_subjects():
    """Preload CSE and AD subjects"""
    existing = {
        (s.code or "", s.branch or "", s.semester)
        for s in Subject.query.all()
    }

    subjects_data = [
        # ---------------- CSE BRANCH ----------------
        # 3rd Semester (CSE)
        ("ES301", "Energy & Environmental Engineering", "CSE", 3),
        ("CS302", "Discrete Structure", "CSE", 3),
        ("CS303", "Data Structure", "CSE", 3),
        ("CS303P", "Data Structure Lab", "CSE", 3),
        ("CS304", "Digital Systems", "CSE", 3),
        ("CS304P", "Digital Systems Lab", "CSE", 3),
        ("CS305", "Object Oriented Programming & Methodology", "CSE", 3),
        ("CS305P", "OOP & Methodology Lab", "CSE", 3),

        # 4th Semester (CSE)
        ("BT401", "Mathematics-III", "CSE", 4),
        ("CS402", "Analysis Design of Algorithm", "CSE", 4),
        ("CS402P", "Analysis Design of Algorithm Lab", "CSE", 4),
        ("CS403", "Software Engineering", "CSE", 4),
        ("CS403P", "Software Engineering Lab", "CSE", 4),
        ("CS404", "Computer Organization & Architecture", "CSE", 4),
        ("CS405", "Operating Systems", "CSE", 4),
        ("CS405P", "Operating Systems Lab", "CSE", 4),

        # 5th Semester (CSE)
        ("CS501", "Theory of Computation", "CSE", 5),
        ("CS502", "Database Management Systems", "CSE", 5),
        ("CS502P", "Database Management Systems Lab", "CSE", 5),
        ("CS503", "Cyber Security", "CSE", 5),
        # Open Electives (5th sem)
        ("CS504A", "Internet and Web Technology", "CSE", 5),
        ("CS504B", "Object Oriented Programming", "CSE", 5),
        ("CS504C", "Introduction to Database Management Systems", "CSE", 5),
        # Python subject + lab
        ("CS506", "Python Programming", "CSE", 5),
        ("CS506P", "Python Programming Lab", "CSE", 5),

        # 6th Semester (CSE)
        ("CS601", "Machine Learning", "CSE", 6),
        ("CS601P", "Machine Learning Lab", "CSE", 6),
        ("CS602", "Computer Networks", "CSE", 6),
        ("CS602P", "Computer Networks Lab", "CSE", 6),
        ("CS603", "Compiler Design", "CSE", 6),
        ("CS603P", "Compiler Design Lab", "CSE", 6),
        # Departmental Elective options
        ("CS603A", "Advanced Computer Architecture", "CSE", 6),
        ("CS603B", "Computer Graphics & Visualization", "CSE", 6),
        # Open Electives (6th sem)
        ("CS604A", "Knowledge Management", "CSE", 6),
        ("CS604B", "Project Management", "CSE", 6),

        # 7th Semester (CSE)
        ("CS701", "Software Architectures", "CSE", 7),
        ("CS702A", "Computational Intelligence", "CSE", 7),
        ("CS702B", "Deep & Reinforcement Learning", "CSE", 7),
        ("CS702C", "Wireless & Mobile Computing", "CSE", 7),
        ("CS702P", "Elective-III Lab (CI / DL & RL / WMC)", "CSE", 7),
        ("CS703A", "Cryptography & Information Security", "CSE", 7),
        ("CS703B", "Data Mining and Warehousing", "CSE", 7),
        ("CS703C", "Agile Software Development", "CSE", 7),
        ("CS703P", "Data Mining and Warehousing Lab", "CSE", 7),
        ("CS704", "Major Project Phase-I", "CSE", 7),
        ("CS705", "Seminar", "CSE", 7),

        # 8th Semester (CSE)
        ("CS802A", "Block Chain Technologies", "CSE", 8),
        ("CS802B", "Cloud Computing", "CSE", 8),
        ("CS802C", "High Performance Computing", "CSE", 8),
        ("CS802D", "Object Oriented Software Engineering", "CSE", 8),
        ("CS802P", "Dept. Elective-IV Lab", "CSE", 8),
        ("CS803A", "Image Processing and Computer Vision", "CSE", 8),
        ("CS803B", "Game Theory with Engineering Applications", "CSE", 8),
        ("CS803C", "Internet of Things", "CSE", 8),
        ("CS803D", "Managing Innovation and Entrepreneurship", "CSE", 8),
        ("CS803P", "Open Elective-IV Lab", "CSE", 8),
        ("CS801", "Internship / Industrial Training", "CSE", 8),
        ("CS804", "Major Project Phase-II", "CSE", 8),

        # ---------------- AD BRANCH (AI & DS) ----------------
        # 3rd Semester (AD)
        ("AD301", "Technical Communication", "AD", 3),
        ("AD302", "Probability and Statistics for Data Science", "AD", 3),
        ("AD303", "Data Structures", "AD", 3),
        ("AD303P", "Data Structures Lab", "AD", 3),
        ("AD304", "Artificial Intelligence", "AD", 3),
        ("AD304P", "Artificial Intelligence Lab", "AD", 3),

        # 4th Semester (AD)
        ("BT401", "Mathematics-III", "AD", 4),
        ("AD402", "Database Management System", "AD", 4),
        ("AD402P", "Database Management System Lab", "AD", 4),
        ("AD403", "Software Engineering with Agile Methodology", "AD", 4),
        ("AD403P", "Software Engineering Lab", "AD", 4),
        ("AD404", "Data Mining", "AD", 4),
        ("AD404P", "Data Mining Lab", "AD", 4),

        # 5th Semester (AD)
        ("AD501", "Theory of Computation", "AD", 5),
        ("AD502", "Machine Learning", "AD", 5),
        ("AD502P", "Machine Learning Lab", "AD", 5),
        ("AD503A", "Internet and Web Technology", "AD", 5),
        ("AD503AP", "Internet and Web Technology Lab", "AD", 5),
        ("AD503B", "Computer Graphics & Multimedia", "AD", 5),
        ("AD503BP", "Computer Graphics & Multimedia Lab", "AD", 5),

        # 6th Semester (AD)
        ("AD601", "Deep Learning", "AD", 6),
        ("AD601P", "Deep Learning Lab", "AD", 6),
        ("AD602", "Computer Networks", "AD", 6),
        ("AD602P", "Computer Networks Lab", "AD", 6),
        ("AD603A", "Data Mining and Warehousing", "AD", 6),
        ("AD603AP", "Data Mining and Warehousing Lab", "AD", 6),
        ("AD603B", "Digital Image Processing", "AD", 6),
        ("AD603BP", "Digital Image Processing Lab", "AD", 6),

        # 7th Semester (AD)
        ("AD701", "AI for Computer Vision", "AD", 7),
        ("AD701P", "AI for Computer Vision Lab", "AD", 7),
        ("AD702A", "Cloud Computing", "AD", 7),
        ("AD702B", "Business Intelligence", "AD", 7),
        ("AD702C", "Computational Intelligence", "AD", 7),
        ("AD702D", "Predictive Analytics", "AD", 7),
        ("AD703", "Seminar", "AD", 7),
        ("AD704", "Major Project Phase-I", "AD", 7),

        # 8th Semester (AD)
        ("AD801", "Big Data", "AD", 8),
        ("AD802A", "Natural Language Processing", "AD", 8),
        ("AD802B", "Reinforcement Learning", "AD", 8),
        ("AD802C", "Robotic Process Automation", "AD", 8),
        ("AD803", "Internship / Industrial Training", "AD", 8),
        ("AD804", "Major Project Phase-II", "AD", 8),
    ]
    added = 0
    for code, name, branch, sem in subjects_data:
        key = (code, branch, sem)
        if key in existing:
            continue

        subject = Subject(
            code=code,
            name=name,
            branch=branch,
            semester=sem,
            is_active=True,
        )
        db.session.add(subject)
        added += 1

    if added > 0:
        db.session.commit()
        print(f"[OK] Subjects preloaded: {added} new subjects added")
    else:
        print(f" All subjects already present")

    return added


def load_students_from_files(data_dir="data"):
    """Load students from Excel/CSV files"""
    os.makedirs(data_dir, exist_ok=True)
    files = [f for f in os.listdir(data_dir) if
             f.lower().endswith(('.xlsx', '.xls', '.csv')) and not f.startswith('~$')]
    total_processed = 0

    if not files:
        print(" No student files found in data directory")
        return

    print("[UPDATE] Syncing students data from files...")

    for fname in files:
        path = os.path.join(data_dir, fname)
        low = fname.lower()

        if '2nd' in low or low.startswith('2') or 'second' in low:
            file_year = 2
        elif '3rd' in low or '3' in low or 'third' in low:
            file_year = 3
        elif '4th' in low or '4' in low or 'fourth' in low:
            file_year = 4
        else:
            file_year = 3

        try:
            if path.lower().endswith(('.xls', '.xlsx')):
                df = pd.read_excel(path)
            else:
                df = pd.read_csv(path)

            df = df.fillna('')
            file_processed = 0

            for _, row in df.iterrows():
                roll_val = row.get('Roll') or row.get('roll') or row.get('Roll No') or row.get('roll_no') or ''
                roll = str(roll_val).strip()

                name_val = row.get('Name') or row.get('name') or row.get('Student Name') or row.get(
                    'student_name') or ''
                name = str(name_val).strip()

                branch_val = row.get('Branch') or row.get('branch') or row.get('Department') or row.get(
                    'department') or 'CSE'
                branch = str(branch_val).strip()

                if roll and name:
                    existing = Student.query.filter_by(roll=roll).first()
                    if existing:
                        if existing.year != file_year or existing.name != name or existing.branch != branch:
                            existing.year = file_year
                            existing.name = name
                            existing.branch = branch
                            file_processed += 1
                    else:
                        new_student = Student(roll=roll, name=name, branch=branch, year=file_year)
                        db.session.add(new_student)
                        file_processed += 1

            db.session.commit()
            print(f" {fname}: Synced {file_processed} students (Year: {file_year})")
            total_processed += file_processed

        except Exception as e:
            print(f"[ERROR] Error processing {fname}: {str(e)}")
            db.session.rollback()

    print(f"[OK] TOTAL STUDENTS SYNCED: {total_processed}")


def activate_all_subjects():
    """Activate ALL subjects in the database"""
    with app.app_context():
        subjects = Subject.query.all()
        activated_count = 0

        for subject in subjects:
            if not subject.is_active:
                subject.is_active = True
                activated_count += 1
                print(f"[OK] Activated: {subject.code} - {subject.name}")

        if activated_count > 0:
            db.session.commit()
            print(f" Activated {activated_count} subjects!")
        else:
            print(" All subjects are already active")

        active_count = Subject.query.filter_by(is_active=True).count()
        total_count = Subject.query.count()
        print(f"[INFO] Subjects Status: {active_count}/{total_count} active")

        return activated_count

def migrate_test_system():
    """Migrate to new test system"""
    with app.app_context():
        try:
            with db.engine.connect() as conn:
                from sqlalchemy import inspect
                inspector = inspect(db.engine)
                existing_columns = [col['name'] for col in inspector.get_columns('tests')]

                new_columns = [
                    "available_from", "available_until", "auto_submit", "prevent_tab_switch", "allow_retake"
                ]

                for col in new_columns:
                    if col not in existing_columns:
                        if col in ['available_from', 'available_until']:
                            conn.execute(text(f"ALTER TABLE tests ADD COLUMN {col} DATETIME"))
                        elif col in ['auto_submit', 'prevent_tab_switch', 'allow_retake']:
                            conn.execute(text(f"ALTER TABLE tests ADD COLUMN {col} BOOLEAN DEFAULT TRUE"))

                conn.execute(text(
                    "UPDATE tests SET available_from = start_time, available_until = end_time WHERE available_from IS NULL"))
                conn.commit()

            print("[OK] Test system migrated successfully!")

        except Exception as e:
            print(f"[ERROR] Migration error: {e}")


def initialize_current_semester():
    """Initialize current semester for all branches and years"""
    branches = ['CSE', 'AD']
    years = [1, 2, 3, 4]

    for branch in branches:
        for year in years:
            existing = CurrentSemester.query.filter_by(
                branch=branch,
                year=year,
                is_active=True
            ).first()

            if not existing:
                current_semester = CurrentSemester(
                    branch=branch,
                    year=year,
                    semester_type='odd',
                    academic_year=get_current_academic_year(),
                    is_active=True
                )
                db.session.add(current_semester)

    db.session.commit()
    print("[OK] Current semester initialized for all branches and years")


def initialize_rgpv_scheme_complete():
    """Initialize complete RGPV scheme data for all semesters"""
    try:
        RGPVScheme.query.delete()
        rgpv_schemes = [
            # CSE Branch - 3rd to 8th semesters
            # Format: (code, name, branch, year, semester, lectures, tutorials, practicals, credits)
            ("CS301", "Energy & Environmental Engineering", "CSE", 2, 3, 3, 1, 0, 4),
            ("CS302", "Discrete Structure", "CSE", 2, 3, 3, 1, 0, 4),
            ("CS303", "Data Structure", "CSE", 2, 3, 3, 0, 2, 4),
            ("CS304", "Digital Systems", "CSE", 2, 3, 3, 0, 2, 4),
            ("CS305", "Object Oriented Programming & Methodology", "CSE", 2, 3, 3, 0, 2, 4),
            ("CS306", "Computer Workshop", "CSE", 2, 3, 0, 0, 4, 2),

            # ---------- 4th Semester (CSE) ----------
            ("BT401", "Mathematics III", "CSE", 2, 4, 3, 1, 0, 4),
            ("CS402", "Analysis Design of Algorithm", "CSE", 2, 4, 2, 1, 2, 4),
            ("CS403", "Software Engineering", "CSE", 2, 4, 3, 1, 2, 5),
            ("CS404", "Computer Organization & Architecture", "CSE", 2, 4, 3, 1, 2, 5),
            ("CS405", "Operating Systems", "CSE", 2, 4, 3, 0, 2, 4),
            ("CS406", "Programming Practices", "CSE", 2, 4, 0, 0, 4, 2),

            # ---------- 5th Semester (CSE) ----------
            ("CS501", "Theory of Computation", "CSE", 3, 5, 3, 0, 2, 4),
            ("CS502", "Database Management Systems", "CSE", 3, 5, 3, 0, 2, 4),
            ("CS503", "Departmental Elective", "CSE", 3, 5, 3, 0, 0, 4),
            ("CS504", "Open Elective", "CSE", 3, 5, 3, 0, 0, 3),
            ("CS505", "Lab (Linux)", "CSE", 3, 5, 0, 0, 4, 2),
            ("CS506", "Lab (Python)", "CSE", 3, 5, 0, 0, 4, 2),

            # ---------- 6th Semester (CSE) ----------
            ("CS601", "Machine Learning", "CSE", 3, 6, 2, 1, 2, 4),
            ("CS602", "Computer Networks", "CSE", 3, 6, 2, 1, 2, 4),
            ("CS603", "Departmental Elective", "CSE", 3, 6, 4, 0, 0, 4),
            ("CS604", "Open Elective", "CSE", 3, 6, 4, 0, 0, 4),
            ("CS605", "Data Analytics Lab", "CSE", 3, 6, 0, 0, 6, 3),
            ("CS606", "Skill Development Lab", "CSE", 3, 6, 0, 0, 6, 3),

            # ---------- 7th Semester (CSE) ----------
            ("CS701", "Software Architectures", "CSE", 4, 7, 2, 1, 2, 4),
            ("CS702", "Departmental Elective", "CSE", 4, 7, 3, 1, 0, 4),
            ("CS703", "Open Elective", "CSE", 4, 7, 3, 0, 0, 3),
            ("CS704", "Departmental Elective Lab", "CSE", 4, 7, 0, 0, 6, 3),
            ("CS705", "Open Elective Lab", "CSE", 4, 7, 0, 0, 6, 3),
            ("CS706", "Major Project-I", "CSE", 4, 7, 0, 0, 8, 4),

            # ---------- 8th Semester (CSE) ----------
            ("CS801", "Internet of Things", "CSE", 4, 8, 2, 1, 2, 4),
            ("CS802", "Departmental Elective", "CSE", 4, 8, 3, 1, 0, 4),
            ("CS803", "Open Elective", "CSE", 4, 8, 3, 0, 0, 3),
            ("CS804", "D/O Elective Lab", "CSE", 4, 8, 0, 0, 6, 3),
            ("CS805", "Major Project-II", "CSE", 4, 8, 0, 0, 8, 4),

            # ======================= AD BRANCH =======================
            # ---------- 3rd Semester (AD) ----------
            ("AD301", "Technical Communication", "AD", 2, 3, 3, 1, 0, 4),
            ("AD302", "Probability and Statistics for Data Science", "AD", 2, 3, 3, 1, 0, 4),
            ("AD303", "Data Structures", "AD", 2, 3, 3, 0, 2, 4),
            ("AD304", "Artificial Intelligence", "AD", 2, 3, 3, 0, 2, 4),
            ("AD305", "Object Oriented Programming & Methodology", "AD", 2, 3, 3, 0, 2, 4),
            ("AD306", "Computer Workshop/Introduction to Python", "AD", 2, 3, 0, 0, 4, 2),

            # ---------- 4th Semester (AD) ----------
            ("BT401", "Mathematics III", "AD", 2, 4, 3, 1, 0, 4),
            ("AD402", "Database Management Systems", "AD", 2, 4, 4, 0, 2, 5),
            ("AD403", "Software Engineering with Agile Methodology", "AD", 2, 4, 4, 0, 2, 5),
            ("AD404", "Data Science", "AD", 2, 4, 3, 0, 2, 4),
            ("AD405", "Operating Systems", "AD", 2, 4, 3, 0, 2, 4),
            ("AD406", "Data Analytics using tools", "AD", 2, 4, 0, 0, 4, 2),

            # ---------- 5th Semester (AD) ----------
            ("AD501", "Theory of Computation", "AD", 3, 5, 3, 0, 2, 4),
            ("AD502", "Machine Learning", "AD", 3, 5, 3, 0, 2, 4),
            ("AD503", "Departmental Elective", "AD", 3, 5, 3, 1, 0, 4),
            ("AD504", "Open Elective", "AD", 3, 5, 3, 0, 0, 3),
            ("AD505", "Departmental Elective Lab", "AD", 3, 5, 0, 0, 4, 2),
            ("AD506", "Linux Lab", "AD", 3, 5, 0, 0, 4, 2),

            # ---------- 6th Semester (AD) ----------
            ("AD601", "Deep Learning", "AD", 3, 6, 2, 1, 2, 4),
            ("AD602", "Computer Networks", "AD", 3, 6, 2, 1, 2, 4),
            ("AD603", "Departmental Elective", "AD", 3, 6, 4, 0, 0, 4),
            ("AD604", "Open Elective", "AD", 3, 6, 4, 0, 0, 4),
            ("AD605", "Departmental Elective Lab", "AD", 3, 6, 0, 0, 6, 3),
            ("AD606", "Open Elective Lab", "AD", 3, 6, 0, 0, 6, 3),

            # ---------- 7th Semester (AD) ----------
            ("AD701", "AI for Computer Vision", "AD", 4, 7, 2, 1, 2, 4),
            ("AD702", "Departmental Elective", "AD", 4, 7, 3, 1, 0, 4),
            ("AD703", "Open Elective", "AD", 4, 7, 3, 0, 0, 3),
            ("AD704", "Departmental Elective Lab", "AD", 4, 7, 0, 0, 6, 3),
            ("AD705", "Open Elective Lab", "AD", 4, 7, 0, 0, 6, 3),
            ("AD706", "Major Project-I", "AD", 4, 7, 0, 0, 8, 4),

            # ---------- 8th Semester (AD) ----------
            ("AD801", "Big Data", "AD", 4, 8, 2, 1, 2, 4),
            ("AD802", "Departmental Elective", "AD", 4, 8, 3, 1, 0, 4),
            ("AD803", "Open Elective", "AD", 4, 8, 3, 0, 0, 3),
            ("AD804", "Departmental/Open Elective Lab", "AD", 4, 8, 0, 0, 6, 3),
            ("AD805", "Major Project-II", "AD", 4, 8, 0, 0, 8, 4),
        ]

        added_count = 0
        for scheme_data in rgpv_schemes:
            try:
                # Unpack the data - ensure exactly 9 values
                if len(scheme_data) == 9:
                    code, name, branch, year, semester, lectures, tutorials, practicals, credits = scheme_data

                    subject = Subject.query.filter_by(
                        code=code,
                        branch=branch,
                        semester=semester
                    ).first()

                    if subject:
                        scheme = RGPVScheme(
                            branch=branch,
                            year=year,
                            semester=semester,
                            subject_id=subject.id,
                            lectures_per_week=lectures,
                            tutorials_per_week=tutorials,
                            practicals_per_week=practicals,
                            credits=credits
                        )
                        db.session.add(scheme)
                        added_count += 1
                else:
                    print(f"⚠️  Skipping invalid scheme data: {scheme_data}")

            except Exception as e:
                print(f"⚠️  Error processing scheme {scheme_data}: {e}")
                continue

        db.session.commit()
        print(f"[OK] RGPV Scheme initialized: {added_count} subjects")

    except Exception as e:
        print(f"[ERROR] RGPV Scheme initialization failed: {e}")
        db.session.rollback()

def get_today_attendance_summary():
    """Get today's attendance summary for all subjects"""
    try:
        today = date.today()

        # Get all subjects that have attendance records today
        subject_ids = db.session.query(Attendance.subject_id).filter(
            Attendance.date == today
        ).distinct().all()

        stats = []
        for (subject_id,) in subject_ids:
            subject = Subject.query.get(subject_id)
            if not subject:
                continue

            # Get professor for this subject
            allotment = ProfessorSubject.query.filter_by(subject_id=subject_id).first()
            professor = User.query.get(allotment.professor_id) if allotment else None

            # Count present and absent students
            present_count = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today,
                status='present'
            ).count()

            absent_count = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today,
                status='absent'
            ).count()

            total_students = present_count + absent_count
            percentage = round((present_count / total_students) * 100, 2) if total_students > 0 else 0

            # Get last updated time
            latest_record = Attendance.query.filter_by(
                subject_id=subject_id,
                date=today
            ).order_by(Attendance.created_at.desc()).first()

            last_updated = latest_record.created_at.strftime('%H:%M') if latest_record else 'N/A'

            stats.append({
                'subject_id': subject_id,
                'subject_code': subject.code,
                'subject_name': subject.name,
                'professor_name': professor.fullname if professor else 'Not Allotted',
                'branch': subject.branch,
                'year': (subject.semester + 1) // 2,  # Calculate year from semester
                'semester': subject.semester,
                'present_count': present_count,
                'absent_count': absent_count,
                'total_students': total_students,
                'percentage': percentage,
                'last_updated': last_updated
            })

        # Sort by last updated (most recent first)
        stats.sort(key=lambda x: x['last_updated'], reverse=True)
        return stats

    except Exception as e:
        print(f"Error in get_today_attendance_summary: {e}")
        return []
def ensure_student_accounts():
    """Create user accounts for all students with default passwords"""
    students = Student.query.filter(Student.roll.isnot(None)).all()
    created_count = 0
    skipped_count = 0

    for student in students:
        if not student.roll or student.roll.strip() == "":
            print(f" Skipped student with missing roll: {student.name}")
            skipped_count += 1
            continue

        existing_user = User.query.filter_by(student_roll=student.roll).first()
        if existing_user:
            skipped_count += 1
            continue

        student_user = User(
            username=student.roll.upper(),
            fullname=student.name,
            email=f"{student.roll.lower()}@college.com",
            role='student',
            branch=student.branch,
            student_roll=student.roll.upper(),
            email_verified=True
        )
        student_user.set_password(student.roll.upper())
        db.session.add(student_user)
        created_count += 1

    db.session.commit()
    print(f"[OK] Created {created_count} new student accounts")
    print(f" Skipped {skipped_count} accounts (already exists or missing roll)")
def get_current_academic_year():
    """Get current academic year based on month"""
    today = datetime.now()
    if today.month >= 7:
        return today.year
    else:
        return today.year - 1


# ========== DATABASE INITIALIZATION FOR RENDER ==========
def init_database():
    """Initialize database tables and create default admin user"""

    with app.app_context():
        try:
            print("🚀 Starting database initialization...")

            # Create all tables
            db.create_all()
            print("✓ Database tables created successfully")

            # Create default admin user if not exists
            if not User.query.filter_by(role='admin').first():
                admin = User(
                    username='admin',
                    fullname='Administrator',
                    email='admin@college.com',
                    password_hash=generate_password_hash('admin123'),
                    role='admin',
                    branch='CSE',
                    email_verified=True,
                    is_active=True
                )
                db.session.add(admin)
                print("✓ Default admin user created")

            # Load initial data if needed
            subject_count = Subject.query.count()
            if subject_count == 0:
                print("✓ Loading initial data...")
                preload_subjects()  # This function is now defined above
                load_students_from_files()  # This function is now defined above
                ensure_student_accounts()
                initialize_current_semester()
                initialize_rgpv_scheme_complete()
                migrate_test_system()

            db.session.commit()
            print("✅ Database initialization completed successfully!")

        except Exception as e:
            print(f"❌ Database initialization failed: {str(e)}")
            import traceback
            traceback.print_exc()


# Run database initialization
init_database()
# Global variables
REPORT_DIR = os.path.join(basedir, 'reports')
UPLOAD_FOLDER = os.path.join(basedir, 'uploads')
NOTES_FOLDER = os.path.join(UPLOAD_FOLDER, 'notes')
PROFILE_PHOTOS_FOLDER = os.path.join(UPLOAD_FOLDER, 'profile_photos')

if not IS_VERCEL:
    os.makedirs(REPORT_DIR, exist_ok=True)
    os.makedirs(NOTES_FOLDER, exist_ok=True)
    os.makedirs(PROFILE_PHOTOS_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'ppt', 'pptx', 'txt', 'jpg', 'png', 'jpeg', 'gif'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

# ========== TIMETABLE CONFIGURATION ==========
COLLEGE_TIMINGS = {
    1: "10:00 - 11:00",
    2: "11:00 - 12:00",
    3: "12:30 - 13:25",
    4: "13:25 - 14:20",
    5: "14:35 - 15:35",
    6: "15:35 - 16:15"
}

LAB_COMBINATIONS = [
    [1, 2],  # P1+P2 = 10:00-12:00
    [3, 4],  # P3+P4 = 12:30-14:20
    [5, 6]  # P5+P6 = 14:35-16:15
]

WORKING_DAYS = {
    1: "Monday",
    2: "Tuesday",
    3: "Wednesday",
    4: "Thursday",
    5: "Friday"
}

COMMON_SUBJECTS_MAP = {
    "Data Structures": {"CSE": "CS303", "AD": "AD303", "is_lab": False, "weekly_slots": 3},
    "Data Structure Lab": {"CSE": "CS303P", "AD": "AD303P", "is_lab": True, "weekly_slots": 1},
    "Object Oriented Programming": {"CSE": "CS305", "AD": "AD305", "is_lab": False, "weekly_slots": 3},
    "OOP Lab": {"CSE": "CS305P", "AD": "AD305P", "is_lab": True, "weekly_slots": 1},
    "Digital Systems": {"CSE": "CS304", "AD": "AD304", "is_lab": False, "weekly_slots": 2},
    "Digital Systems Lab": {"CSE": "CS304P", "AD": "AD304P", "is_lab": True, "weekly_slots": 1},
    "Mathematics-III": {"CSE": "BT401", "AD": "BT401", "is_lab": False, "weekly_slots": 3},
    "Theory of Computation": {"CSE": "CS501", "AD": "AD501", "is_lab": False, "weekly_slots": 2},
    "Database Management Systems": {"CSE": "CS502", "AD": "AD402", "is_lab": False, "weekly_slots": 2},
    "DBMS Lab": {"CSE": "CS502P", "AD": "AD402P", "is_lab": True, "weekly_slots": 1},
    "Machine Learning": {"CSE": "CS601", "AD": "AD502", "is_lab": False, "weekly_slots": 2},
    "Machine Learning Lab": {"CSE": "CS601P", "AD": "AD502P", "is_lab": True, "weekly_slots": 1},
    "Computer Networks": {"CSE": "CS602", "AD": "AD602", "is_lab": False, "weekly_slots": 2},
    "Computer Networks Lab": {"CSE": "CS602P", "AD": "AD602P", "is_lab": True, "weekly_slots": 1},
    "Deep Learning": {"CSE": "CS601", "AD": "AD601", "is_lab": False, "weekly_slots": 2},
    "Deep Learning Lab": {"CSE": "CS601P", "AD": "AD601P", "is_lab": True, "weekly_slots": 1}
}


# ========== HELPER FUNCTIONS ==========
def get_student_subject_attendance(student_id, subject_id, start_date, end_date):
    """Get student's attendance for a specific subject in date range"""
    present_count = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.subject_id == subject_id,
        Attendance.date.between(start_date, end_date),
        Attendance.status == 'present'
    ).count()

    return {
        'present_count': present_count,
        'student_id': student_id,
        'subject_id': subject_id
    }

def get_current_semester_type():
    """Determine if current semester is odd or even"""
    today = datetime.now()
    if today.month >= 7:
        return 'odd'
    else:
        return 'even'


def get_active_semester_for_branch_year(branch, year):
    """Get currently active semester for specific branch and year"""
    return CurrentSemester.query.filter_by(
        branch=branch,
        year=year,
        is_active=True
    ).first()


def get_working_days_for_month(year, month):
    """Get all working days (Monday-Friday) for a month with dates"""
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    working_days = {}
    day_counter = 1

    current_date = first_day
    while current_date <= last_day:
        if current_date.weekday() <= 4:  # Monday to Friday
            working_days[day_counter] = current_date
            day_counter += 1
        current_date += timedelta(days=1)

    return working_days


def get_professor_for_subject(subject_id):
    """Get professor name for a subject"""
    allotment = ProfessorSubject.query.filter_by(subject_id=subject_id).first()
    if allotment:
        professor = User.query.get(allotment.professor_id)
        return professor.fullname if professor else "Not Allotted"
    return "Not Allotted"


def get_semesters_for_branch_year(branch, year):
    """Get which semesters should be active based on current period"""
    current_semester = get_active_semester_for_branch_year(branch, year)
    if not current_semester:
        return []

    if current_semester.semester_type == 'odd':
        odd_semester_mapping = {1: [1], 2: [3], 3: [5], 4: [7]}
        return odd_semester_mapping.get(year, [])
    else:
        even_semester_mapping = {1: [2], 2: [4], 3: [6], 4: [8]}
        return even_semester_mapping.get(year, [])


def get_student_current_semesters(branch, year):
    """Get current semesters for a student based on their branch and year"""
    return get_semesters_for_branch_year(branch, year)


def get_all_active_semesters():
    """Get all active semester configurations"""
    return CurrentSemester.query.filter_by(is_active=True).all()


def generate_otp(length=6):
    """Generate numeric OTP"""
    return ''.join(random.choices('0123456789', k=length))


def year_word(year_no):
    """Convert year number to word"""
    mapping = {1: "First", 2: "Second", 3: "Third", 4: "Fourth"}
    return mapping.get(year_no, f"Year {year_no}")


def get_year_word(year_no):
    """Convert year number to word"""
    return year_word(year_no)


def get_active_subjects(branch, year, semester=None):
    """Get only active subjects that are actually taught"""
    if semester:
        return Subject.query.filter_by(
            branch=branch,
            semester=semester,
            is_active=True
        ).order_by(Subject.code).all()
    else:
        sem1 = year * 2 - 1
        sem2 = year * 2
        return Subject.query.filter(
            Subject.branch == branch,
            Subject.semester.in_([sem1, sem2]),
            Subject.is_active == True
        ).order_by(Subject.semester, Subject.code).all()


def allowed_file(filename):
    """Check if file extension is allowed"""
    if '.' not in filename:
        return False
    file_ext = filename.rsplit('.', 1)[1].lower()
    return file_ext in ALLOWED_EXTENSIONS


def save_profile_photo(file, user_id):
    """Save profile photo and return filename"""
    if file and allowed_file(file.filename):
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"user_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_ext}"
        file_path = os.path.join(PROFILE_PHOTOS_FOLDER, filename)
        file.save(file_path)
        return filename
    return None


def has_unseen_notices(user_id, user_role, branch=None, year=None):
    """Check if user has unseen notices"""
    from datetime import datetime
    notices = Notice.query.filter(
        Notice.is_active == True,
        Notice.expires_at >= datetime.now()
    ).all()

    relevant_notices = []
    for notice in notices:
        if is_notice_relevant(notice, user_role, branch, year):
            if str(user_id) not in (notice.seen_by or '').split(','):
                relevant_notices.append(notice)

    return len(relevant_notices) > 0


def is_notice_relevant(notice, user_role, branch=None, year=None):
    """Check if notice is relevant for this user"""
    if notice.target_audience == 'all':
        return True
    elif notice.target_audience == 'students' and user_role == 'student':
        return True
    elif notice.target_audience == 'professors' and user_role == 'professor':
        return True
    elif notice.target_audience == 'students' and user_role == 'student':
        if notice.branch and branch != notice.branch:
            return False
        if notice.year and year != notice.year:
            return False
        return True
    return False


def mark_notice_as_seen(notice_id, user_id):
    """Mark a notice as seen by user"""
    notice = Notice.query.get(notice_id)
    if notice:
        seen_users = (notice.seen_by or '').split(',')
        if str(user_id) not in seen_users:
            seen_users.append(str(user_id))
            notice.seen_by = ','.join(filter(None, seen_users))
            db.session.commit()


# ========== SMART TEST SYSTEM FUNCTIONS ==========
def auto_submit_test(attempt_id):
    """Automatically submit test when time ends"""
    attempt = TestAttempt.query.get_or_404(attempt_id)
    if not attempt.submitted:
        answers = StudentAnswer.query.filter_by(attempt_id=attempt_id).all()
        total_marks = sum(answer.marks_obtained for answer in answers)

        attempt.end_time = datetime.now()
        attempt.total_marks_obtained = total_marks
        attempt.submitted = True
        db.session.commit()
        print(f"[OK] Test auto-submitted for attempt {attempt_id}")


# ========== TIMETABLE FUNCTIONS ==========
def get_professor_for_common_subject(cse_subject_id, ad_subject_id):
    """Get professor for common subject with proper error handling"""
    try:
        cse_allotment = ProfessorSubject.query.filter_by(subject_id=cse_subject_id).first()
        if cse_allotment:
            professor = User.query.get(cse_allotment.professor_id)
            if professor:
                return professor

        ad_allotment = ProfessorSubject.query.filter_by(subject_id=ad_subject_id).first()
        if ad_allotment:
            professor = User.query.get(ad_allotment.professor_id)
            if professor:
                return professor

        active_professor = User.query.filter_by(
            role='professor',
            is_active=True
        ).first()

        return active_professor

    except Exception as e:
        print(f"[ERROR] Error getting professor for common subject: {e}")
        return None


def is_slot_available(timetable, day, period):
    """Check if a slot is available in the timetable"""
    try:
        return timetable['slots'][day][period]['subject'] is None
    except Exception as e:
        print(f"[ERROR] Error checking slot availability: {e}")
        return False


def is_professor_available(professor_id, day, period, faculty_daily_periods):
    """Check if professor is available at given day and period"""
    try:
        return period not in faculty_daily_periods[professor_id][day]
    except Exception as e:
        print(f"[ERROR] Error checking professor availability: {e}")
        return False


def get_common_subjects_for_semester(semester):
    """Get common subjects for specific semester with better filtering"""
    common_subjects = {}

    for name, details in COMMON_SUBJECTS_MAP.items():
        cse_subject = Subject.query.filter_by(
            code=details["CSE"],
            semester=semester,
            is_active=True
        ).first()

        ad_subject = Subject.query.filter_by(
            code=details["AD"],
            semester=semester,
            is_active=True
        ).first()

        if cse_subject and ad_subject:
            common_subjects[name] = details
            print(f"[OK] Found common subject: {name} - CSE:{cse_subject.code}, AD:{ad_subject.code}")

    return common_subjects


def allocate_common_subjects_optimized(timetables, branches, years, semesters):
    """Optimized common subject allocation with better slot distribution"""
    print("[UPDATE] Optimized common subjects allocation...")

    for semester in semesters:
        common_subjects = get_common_subjects_for_semester(semester)

        if not common_subjects:
            print(f" No common subjects found for semester {semester}")
            continue

        theory_subjects = {name: details for name, details in common_subjects.items() if not details["is_lab"]}
        lab_subjects = {name: details for name, details in common_subjects.items() if details["is_lab"]}

        print(f" Theory subjects: {len(theory_subjects)}, Lab subjects: {len(lab_subjects)}")

        # Allocate theory subjects
        for common_name, details in theory_subjects.items():
            weekly_slots = details.get("weekly_slots", 3)
            print(f"[UPDATE] Allocating theory: {common_name} ({weekly_slots} slots)")

            cse_subject = Subject.query.filter_by(code=details["CSE"], semester=semester).first()
            ad_subject = Subject.query.filter_by(code=details["AD"], semester=semester).first()
            professor = get_professor_for_common_subject(cse_subject.id, ad_subject.id)

            if not cse_subject or not ad_subject or not professor:
                print(f"[ERROR] Missing data for {common_name}, skipping...")
                continue

            slots_assigned = 0
            attempts = 0
            max_attempts = 200

            days = list(range(1, 6))
            periods = list(range(1, 7))

            while slots_assigned < weekly_slots and attempts < max_attempts:
                attempts += 1
                day = random.choice(days)
                period = random.choice(periods)

                if is_slot_available_for_common(timetables, branches, years[0], semester, day, period):
                    room = f"Room-{random.randint(101, 110)}"

                    for branch in branches:
                        key = f"{branch}_{years[0]}_{semester}"
                        subject = cse_subject if branch == "CSE" else ad_subject

                        timetables[key]['slots'][day][period] = {
                            'subject': subject,
                            'faculty': professor,
                            'room': room,
                            'slot_type': 'lecture',
                            'is_common': True,
                            'common_name': common_name
                        }

                    print(f"  [OK] {common_name} at Day {day}, Period {period}")
                    slots_assigned += 1

            if slots_assigned < weekly_slots:
                print(f"[WARNING] Only allocated {slots_assigned}/{weekly_slots} slots for {common_name}")

        # Allocate lab subjects
        for common_name, details in lab_subjects.items():
            print(f"[UPDATE] Allocating lab: {common_name}")

            cse_subject = Subject.query.filter_by(code=details["CSE"], semester=semester).first()
            ad_subject = Subject.query.filter_by(code=details["AD"], semester=semester).first()
            professor = get_professor_for_common_subject(cse_subject.id, ad_subject.id)

            if not cse_subject or not ad_subject or not professor:
                print(f"[ERROR] Missing data for lab {common_name}, skipping...")
                continue

            lab_allocated = False
            lab_combinations = [[1, 2], [3, 4], [5, 6]]

            for day in range(1, 6):
                for start_period, end_period in lab_combinations:
                    if (is_slot_available_for_common(timetables, branches, years[0], semester, day, start_period) and
                            is_slot_available_for_common(timetables, branches, years[0], semester, day, end_period)):

                        room = f"Lab-{random.randint(1, 5)}"

                        for branch in branches:
                            key = f"{branch}_{years[0]}_{semester}"
                            subject = cse_subject if branch == "CSE" else ad_subject

                            timetables[key]['slots'][day][start_period] = {
                                'subject': subject,
                                'faculty': professor,
                                'room': room,
                                'slot_type': 'lab',
                                'is_common': True,
                                'common_name': common_name
                            }
                            timetables[key]['slots'][day][end_period] = {
                                'subject': subject,
                                'faculty': professor,
                                'room': room,
                                'slot_type': 'lab',
                                'is_common': True,
                                'common_name': common_name
                            }

                        print(f"  [OK] Lab {common_name} at Day {day}, Periods {start_period}-{end_period}")
                        lab_allocated = True
                        break

                if lab_allocated:
                    break

            if not lab_allocated:
                print(f"[ERROR] Could not allocate lab {common_name}")


def is_slot_available_for_common(timetables, branches, year, semester, day, slot):
    """Check if slot is available for all branches"""
    try:
        for branch in branches:
            key = f"{branch}_{year}_{semester}"
            if key not in timetables:
                return False

            slot_data = timetables[key]['slots'][day][slot]
            if slot_data['subject'] is not None:
                return False

        return True
    except Exception as e:
        print(f"[ERROR] Error checking slot availability: {e}")
        return False


def generate_smart_timetable(branches, years, semesters):
    """Optimized timetable generation with better slot management"""
    print(f"[SUCCESS] Generating optimized timetable for branches={branches}, years={years}, semesters={semesters}")

    try:
        professors = User.query.filter_by(role='professor', is_active=True).all()

        if not professors:
            print("[ERROR] No active professors found!")
            return {}

        timetables = {}
        for branch in branches:
            for year in years:
                for semester in semesters:
                    key = f"{branch}_{year}_{semester}"
                    timetable = {
                        'college_name': 'Shri Balaji Institute of Technology & Management, Betul (M.P.)',
                        'department': f'Department of {"Computer Science and Engineering" if branch == "CSE" else "Data Analytics"}',
                        'branch': branch,
                        'year': year,
                        'semester': semester,
                        'timings': COLLEGE_TIMINGS,
                        'days': WORKING_DAYS,
                        'slots': {}
                    }

                    for day_num in WORKING_DAYS.keys():
                        timetable['slots'][day_num] = {}
                        for period in range(1, 7):
                            timetable['slots'][day_num][period] = {
                                'subject': None,
                                'faculty': None,
                                'room': f"Room-{random.randint(101, 110)}",
                                'slot_type': 'lecture'
                            }
                    timetables[key] = timetable

        print("[UPDATE] Step 1: Allocating common subjects...")
        allocate_common_subjects_optimized(timetables, branches, years, semesters)

        print("[UPDATE] Step 2: Allocating remaining subjects...")
        remaining_subjects = []
        for branch in branches:
            for year in years:
                for semester in semesters:
                    subjects_with_professors = db.session.query(Subject, ProfessorSubject).join(
                        ProfessorSubject, Subject.id == ProfessorSubject.subject_id
                    ).filter(
                        Subject.branch == branch,
                        Subject.semester == semester,
                        Subject.is_active == True
                    ).all()

                    for subject, allotment in subjects_with_professors:
                        is_common = False
                        for common_details in COMMON_SUBJECTS_MAP.values():
                            if subject.code in [common_details["CSE"], common_details["AD"]]:
                                is_common = True
                                break

                        if not is_common:
                            professor = User.query.get(allotment.professor_id)
                            if professor:
                                remaining_subjects.append({
                                    'subject': subject,
                                    'professor': professor,
                                    'branch': branch,
                                    'year': year,
                                    'semester': semester
                                })

        print(f"[UPDATE] Found {len(remaining_subjects)} remaining subjects to allocate")

        lab_subjects = [s for s in remaining_subjects if 'Lab' in s['subject'].name or s['subject'].code.endswith('P')]
        theory_subjects = [s for s in remaining_subjects if s not in lab_subjects]
        remaining_subjects_sorted = lab_subjects + theory_subjects

        faculty_daily_periods = {prof.id: {day: set() for day in range(1, 6)} for prof in professors}

        allocated_count = 0
        for subj_data in remaining_subjects_sorted:
            subject = subj_data['subject']
            professor = subj_data['professor']
            branch = subj_data['branch']
            year = subj_data['year']
            semester = subj_data['semester']

            key = f"{branch}_{year}_{semester}"
            if key not in timetables:
                continue

            if subject in lab_subjects:
                required_slots = 1
                slot_type = "lab"
                possible_starts = [1, 3, 5]
            else:
                scheme = RGPVScheme.query.filter_by(
                    branch=branch,
                    year=year,
                    semester=semester,
                    subject_id=subject.id
                ).first()
                required_slots = scheme.lectures_per_week if scheme else 2
                slot_type = "lecture"
                possible_starts = list(range(1, 7))

            slots_assigned = 0
            max_attempts = 100

            while slots_assigned < required_slots and max_attempts > 0:
                max_attempts -= 1
                day = random.choice(range(1, 6))

                if slot_type == "lab":
                    for start_period in possible_starts:
                        end_period = start_period + 1
                        if (is_slot_available(timetables[key], day, start_period) and
                                is_slot_available(timetables[key], day, end_period) and
                                is_professor_available(professor.id, day, start_period, faculty_daily_periods) and
                                is_professor_available(professor.id, day, end_period, faculty_daily_periods)):
                            room = f"Lab-{random.randint(1, 5)}"

                            timetables[key]['slots'][day][start_period] = {
                                'subject': subject,
                                'faculty': professor,
                                'room': room,
                                'slot_type': 'lab'
                            }
                            timetables[key]['slots'][day][end_period] = {
                                'subject': subject,
                                'faculty': professor,
                                'room': room,
                                'slot_type': 'lab'
                            }

                            faculty_daily_periods[professor.id][day].add(start_period)
                            faculty_daily_periods[professor.id][day].add(end_period)
                            slots_assigned += 1
                            allocated_count += 1
                            print(
                                f"[OK] Allocated lab {subject.code} at Day {day}, Periods {start_period}-{end_period}")
                            break

                else:
                    period = random.choice(possible_starts)
                    if (is_slot_available(timetables[key], day, period) and
                            is_professor_available(professor.id, day, period, faculty_daily_periods)):
                        room = f"Room-{random.randint(101, 110)}"
                        timetables[key]['slots'][day][period] = {
                            'subject': subject,
                            'faculty': professor,
                            'room': room,
                            'slot_type': 'lecture'
                        }

                        faculty_daily_periods[professor.id][day].add(period)
                        slots_assigned += 1
                        allocated_count += 1
                        print(f"[OK] Allocated theory {subject.code} at Day {day}, Period {period}")

        print(f"[SUCCESS] Timetable generation completed! Allocated {allocated_count} additional subjects")
        return timetables

    except Exception as e:
        print(f"[ERROR] Error in generate_smart_timetable: {e}")
        import traceback
        traceback.print_exc()
        return {}


def save_timetable_to_db(timetables):
    """Save generated timetable to database with error handling"""
    if not timetables:
        print("[ERROR] No timetables to save")
        return False

    try:
        for key in timetables.keys():
            parts = key.split('_')
            if len(parts) == 3:
                branch, year, semester = parts
                TimetableSlot.query.filter(
                    TimetableSlot.branch == branch,
                    TimetableSlot.year == int(year),
                    TimetableSlot.semester == int(semester)
                ).delete()

        saved_count = 0
        for key, timetable_data in timetables.items():
            parts = key.split('_')
            if len(parts) != 3:
                continue

            branch, year, semester = parts
            for day_num, day_data in timetable_data['slots'].items():
                for period_num, slot_data in day_data.items():
                    if (slot_data['subject'] and slot_data['faculty'] and
                            hasattr(slot_data['subject'], 'id') and
                            hasattr(slot_data['faculty'], 'id')):
                        slot = TimetableSlot(
                            branch=branch,
                            year=int(year),
                            semester=int(semester),
                            day_of_week=day_num,
                            period_number=period_num,
                            subject_id=slot_data['subject'].id,
                            faculty_id=slot_data['faculty'].id,
                            room_number=slot_data['room'],
                            slot_type=slot_data['slot_type'],
                            is_common=slot_data.get('is_common', False),
                            common_name=slot_data.get('common_name', '')
                        )
                        db.session.add(slot)
                        saved_count += 1

        db.session.commit()
        print(f"[OK] Timetable saved successfully! {saved_count} slots saved for {len(timetables)} combinations")
        return True

    except Exception as e:
        print(f"[ERROR] Error saving timetable: {e}")
        db.session.rollback()
        return False


# ========== EMAIL SERVICE ==========
def send_email(to_email, subject, body):
    """Send email using Brevo SMTP - Special setup for Brevo"""
    try:
        print(f"🔄 Attempting to send email to: {to_email}")

        # Create message
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_DEFAULT_SENDER']
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))

        print("🔧 Step 1: Connecting to Brevo SMTP...")

        # Brevo specific setup
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.set_debuglevel(1)  # Detailed debug output

        print("🔧 Step 2: Starting TLS...")
        server.starttls()

        print("🔧 Step 3: Brevo Authentication...")
        # Brevo requires both username and password for authentication
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])

        print("🔧 Step 4: Sending email...")
        text = msg.as_string()
        server.sendmail(app.config['MAIL_DEFAULT_SENDER'], to_email, text)

        print("🔧 Step 5: Closing connection...")
        server.quit()

        # Log successful email
        email_log = EmailLog(
            recipient=to_email,
            subject=subject,
            body=body,
            status='sent'
        )
        db.session.add(email_log)
        db.session.commit()

        print(f"✅ Email sent successfully to {to_email}")
        return True

    except smtplib.SMTPAuthenticationError as e:
        print(f"❌ SMTP Authentication Failed: {e}")
        print("💡 Brevo Auth Tip: Make sure you're using the SMTP key, not API key")
        return False

    except Exception as e:
        print(f"❌ Email sending error: {e}")
        # But still try to send (Brevo might have different behavior)
        return False


def send_otp_via_brevo_api(email, otp):
    """Try simple text email"""
    print("=" * 60)
    print(f"🎯 OTP FOR {email}: {otp}")
    print("=" * 60)

    try:
        import sib_api_v3_sdk
        from sib_api_v3_sdk.rest import ApiException

        configuration = sib_api_v3_sdk.Configuration()
        configuration.api_key['api-key'] = os.getenv('BREVO_API_KEY')

        api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(configuration))

        # 🎯 SIMPLE TEXT EMAIL - NO HTML
        send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(
            to=[{"email": email}],
            text_content=f"Your OTP is: {otp}",
            sender={"name": "College", "email": "sbitmstudy@gmail.com"},
            subject=f"OTP: {otp}"
        )

        api_response = api_instance.send_transac_email(send_smtp_email)
        print(f"✅ Simple email sent! Message ID: {api_response.message_id}")
        return True

    except Exception as e:
        print(f"❌ Brevo failed: {e}")
        return True  # Fallback to console
def send_otp_email(email, otp):
    """Send OTP email using Brevo API"""
    print(f"📧 Sending OTP to: {email}")
    print(f"🔑 OTP: {otp}")

    # Try Brevo Transactional API
    success = send_otp_via_brevo_api(email, otp)

    if success:
        print("✅ OTP sent via Brevo Transactional API")
        return True
    else:
        # Fallback
        print("=" * 50)
        print(f"🎯 OTP FOR {email}: {otp}")
        print("=" * 50)
        return True

# ========== JINJA2 FILTERS & CONTEXT ==========
@app.template_filter('endswith')
def endswith_filter(s, suffix):
    if s is None:
        return False
    return str(s).endswith(str(suffix))


@app.route('/test_simple_email')
def test_simple_email():
    """Test with simple text email"""
    test_email = "ravikumarmohane@gmail.com"
    test_otp = generate_otp()

    print("🧪 Testing SIMPLE text email...")
    success = send_otp_via_brevo_api(test_email, test_otp)

    return f"""
    <h3>Simple Email Test</h3>
    <p>Email: {test_email}</p>
    <p>OTP: {test_otp}</p>
    <p>Result: {'✅ SUCCESS' if success else '❌ FAILED'}</p>
    <p>Check server console for detailed logs</p>
    """
@app.template_filter('startswith')
def startswith_filter(s, prefix):
    if s is None:
        return False
    return str(s).startswith(str(prefix))


@app.template_filter('contains')
def contains_filter(s, substring):
    if s is None:
        return False
    return str(substring) in str(s)


@app.context_processor
def utility_processor():
    from datetime import datetime, timedelta
    from flask import request
    import os

    def today_minus_7_days():
        return datetime.now() - timedelta(days=7)

    def get_today_attendance_count(subject_id):
        try:
            return Attendance.query.filter_by(
                subject_id=subject_id,
                date=date.today()
            ).count()
        except Exception as e:
            print(f"Error in get_today_attendance_count: {e}")
            return 0

    def get_total_classes_count(subject_id):
        try:
            return db.session.query(db.func.count(db.func.distinct(Attendance.date))).filter_by(
                subject_id=subject_id
            ).scalar() or 0
        except Exception as e:
            print(f"Error in get_total_classes_count: {e}")
            return 0

    def get_student_count(branch, year):
        try:
            return Student.query.filter_by(
                branch=branch,
                year=year
            ).count()
        except Exception as e:
            print(f"Error in get_student_count: {e}")
            return 0

    def get_today_attendance_summary():
        """Get today's attendance summary for all subjects"""
        try:
            today = date.today()

            # Get all subjects that have attendance records today
            subject_ids = db.session.query(Attendance.subject_id).filter(
                Attendance.date == today
            ).distinct().all()

            stats = []
            for (subject_id,) in subject_ids:
                subject = Subject.query.get(subject_id)
                if not subject:
                    continue

                # Get professor for this subject
                allotment = ProfessorSubject.query.filter_by(subject_id=subject_id).first()
                professor = User.query.get(allotment.professor_id) if allotment else None

                # Count present and absent students
                present_count = Attendance.query.filter_by(
                    subject_id=subject_id,
                    date=today,
                    status='present'
                ).count()

                absent_count = Attendance.query.filter_by(
                    subject_id=subject_id,
                    date=today,
                    status='absent'
                ).count()

                total_students = present_count + absent_count
                percentage = round((present_count / total_students) * 100, 2) if total_students > 0 else 0

                # Get last updated time
                latest_record = Attendance.query.filter_by(
                    subject_id=subject_id,
                    date=today
                ).order_by(Attendance.created_at.desc()).first()

                last_updated = latest_record.created_at.strftime('%H:%M') if latest_record else 'N/A'

                stats.append({
                    'subject_id': subject_id,
                    'subject_code': subject.code,
                    'subject_name': subject.name,
                    'professor_name': professor.fullname if professor else 'Not Allotted',
                    'branch': subject.branch,
                    'year': (subject.semester + 1) // 2,  # Calculate year from semester
                    'semester': subject.semester,
                    'present_count': present_count,
                    'absent_count': absent_count,
                    'total_students': total_students,
                    'percentage': percentage,
                    'last_updated': last_updated
                })

            # Sort by last updated (most recent first)
            stats.sort(key=lambda x: x['last_updated'], reverse=True)
            return stats

        except Exception as e:
            print(f"Error in get_today_attendance_summary: {e}")
            return []

    def get_today_date():
        return date.today().strftime('%d %b %Y')

    def has_unseen_notices_context(user_id, user_role, branch=None, year=None):
        try:
            return has_unseen_notices(user_id, user_role, branch, year)
        except Exception as e:
            print(f"Error in has_unseen_notices_context: {e}")
            return False

    # Additional useful functions for templates
    def get_current_year():
        return datetime.now().year

    def get_current_month():
        return datetime.now().strftime('%B')

    def format_datetime(dt, format_str='%d %b %Y %H:%M'):
        """Format datetime for templates"""
        if dt:
            return dt.strftime(format_str)
        return ''

    def is_active_route(route_name):
        """Check if current route is active for navigation highlighting"""
        return request.endpoint == route_name

    def get_app_name():
        """Get application name"""
        return "College Attendance System"

    def get_app_version():
        """Get application version"""
        return "2.0"

    return dict(
        # Basic Python functions
        enumerate=enumerate,
        len=len,
        str=str,
        int=int,
        float=float,
        list=list,
        dict=dict,

        # Date/Time functions
        date=date,
        datetime=datetime,
        timedelta=timedelta,

        # Attendance functions
        get_today_attendance_count=get_today_attendance_count,
        get_total_classes_count=get_total_classes_count,
        get_student_count=get_student_count,
        get_today_attendance_summary=get_today_attendance_summary,

        # Utility functions
        get_today_date=get_today_date,
        today_minus_7_days=today_minus_7_days,
        has_unseen_notices=has_unseen_notices_context,
        get_current_year=get_current_year,
        get_current_month=get_current_month,
        format_datetime=format_datetime,
        is_active_route=is_active_route,
        get_app_name=get_app_name,
        get_app_version=get_app_version,

        # Additional useful template functions
        round=round,
        zip=zip,
        range=range
    )
# ========== AUTHENTICATION ROUTES ==========
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'professor':
            return redirect(url_for('prof_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'professor':
            return redirect(url_for('prof_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))

    if request.method == 'POST':
        login_input = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        print(f"[DEBUG] Login attempt for: '{login_input}'")

        user = User.query.filter(db.func.lower(User.email) == login_input.lower()).first()
        if not user:
            user = User.query.filter(db.func.lower(User.username) == login_input.lower()).first()

        if user:
            print(f"[DEBUG] User found: {user.username} (Role: {user.role})")
            if user.check_password(password):
                print(f"[DEBUG] Password correct for {user.username}")
                if not user.email_verified and user.role != 'admin' and user.role != 'student':
                    flash('Please verify your email before logging in', 'warning')
                    return redirect(url_for('login'))

                login_user(user)
                flash(f'Welcome back, {user.fullname}!', 'success')

                if user.role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                elif user.role == 'professor':
                    return redirect(url_for('prof_dashboard'))
                else:
                    return redirect(url_for('student_dashboard'))
            else:
                print(f"[DEBUG] Password INCORRECT for {user.username}")
                flash('Invalid email/username or password', 'danger')
        else:
            print(f"[DEBUG] No user found for: '{login_input}'")
            flash('Invalid email/username or password', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        print(f"🔑 Password reset requested for: {email}")

        user = User.query.filter_by(email=email).first()

        if user:
            print(f"✅ User found: {user.fullname}")

            otp = generate_otp()
            expires = datetime.now() + timedelta(minutes=10)

            # Delete existing OTPs
            PasswordResetOTP.query.filter_by(user_id=user.id, used=False).delete()

            reset_entry = PasswordResetOTP(
                user_id=user.id,
                otp_code=otp,
                expires_at=expires,
                used=False
            )
            db.session.add(reset_entry)
            db.session.commit()

            print(f"📧 Sending OTP via Brevo API...")

            # Send OTP
            if send_otp_email(user.email, otp):
                flash('OTP has been sent to your email! Please check your inbox.', 'success')
            else:
                flash('Failed to send OTP. Please try again.', 'danger')

            return redirect(url_for('reset_password', email=email))
        else:
            print(f"❌ User not found: {email}")
            flash('Email address not found. Please check your email.', 'danger')

    return render_template('auth/forgot_password.html')
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    email = request.args.get('email', '').strip()

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        otp = request.form.get('otp', '').strip()
        new_password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not all([email, otp, new_password, confirm_password]):
            flash('All fields are required', 'warning')
            return redirect(url_for('reset_password', email=email))

        if new_password != confirm_password:
            flash('Passwords do not match', 'danger')
            return redirect(url_for('reset_password', email=email))

        user = User.query.filter_by(email=email).first()
        if not user:
            flash('Invalid email or OTP', 'danger')
            return redirect(url_for('reset_password'))

        now = datetime.now()

        otp_entry = PasswordResetOTP.query.filter(
            PasswordResetOTP.user_id == user.id,
            PasswordResetOTP.otp_code == otp,
            PasswordResetOTP.used == False,
            PasswordResetOTP.expires_at >= now
        ).first()

        if not otp_entry:
            flash('Invalid or expired OTP', 'danger')
            return redirect(url_for('reset_password', email=email))

        user.set_password(new_password)
        otp_entry.used = True
        db.session.commit()

        flash('Password has been reset successfully. Please login.', 'success')
        return redirect(url_for('login'))

    return render_template('auth/reset_password.html', email=email)


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change password for logged-in users"""
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'danger')
            return redirect(url_for('change_password'))

        if not current_user.check_password(current_password):
            flash('Current password is incorrect', 'danger')
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match', 'danger')
            return redirect(url_for('change_password'))

        if len(new_password) < 6:
            flash('Password must be at least 6 characters long', 'danger')
            return redirect(url_for('change_password'))

        current_user.set_password(new_password)
        db.session.commit()

        flash('Password changed successfully!', 'success')
        return redirect(url_for('student_dashboard' if current_user.role == 'student' else 'prof_dashboard'))

    return render_template('auth/change_password.html')


# ========== ADMIN ROUTES ==========
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    professors = User.query.filter_by(role='professor').all()
    all_subjects = Subject.query.order_by(Subject.branch, Subject.semester, Subject.code).all()

    allotments = ProfessorSubject.query.all()
    allotments_info = []
    allotted_subject_ids = set()

    for a in allotments:
        prof = User.query.get(a.professor_id)
        subj = Subject.query.get(a.subject_id)
        if prof and subj:
            allotments_info.append({
                "id": a.id,
                "prof": prof,
                "subj": subj
            })
            allotted_subject_ids.add(subj.id)

    active_subjects = [s for s in all_subjects if s.is_active]
    available_subjects = [s for s in active_subjects if s.id not in allotted_subject_ids]

    reports = AttendanceReport.query.order_by(AttendanceReport.date.desc()).limit(50).all()

    current_semester = CurrentSemester.query.filter_by(is_active=True).order_by(
        CurrentSemester.created_at.desc()).first()
    active_semesters = []
    if current_semester:
        active_semesters = get_semesters_for_branch_year(current_semester.branch, current_semester.year)

    return render_template(
        'admin/dashboard.html',
        professors=professors,
        subjects=available_subjects,
        allotments_info=allotments_info,
        reports=reports,
        current_semester=current_semester,
        active_semesters=active_semesters,
        all_subjects=all_subjects
    )
@app.route('/timetable/download/<branch>/<int:year>/<int:semester>')
@login_required
def download_timetable(branch, year, semester):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    timetable = get_timetable_from_db(branch, year, semester)
    if not timetable:
        flash('Timetable not found for selected branch/year/semester', 'warning')
        return redirect(url_for('admin_timetable'))

    import csv
    import io
    from flask import Response

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['Day', 'Period', 'Time', 'Subject', 'Faculty', 'Room'])

    for day_num, day_name in WORKING_DAYS.items():
        for period_num, timing in COLLEGE_TIMINGS.items():
            slot = timetable['slots'][day_num][period_num]

            subject_text = ''
            if slot['subject']:
                subject_text = f"{slot['subject'].code} - {slot['subject'].name}"

            faculty_text = slot['faculty'].fullname if slot['faculty'] else ''
            room_text = slot['room'] or ''

            writer.writerow([
                day_name,
                period_num,
                timing,
                subject_text,
                faculty_text,
                room_text
            ])

    csv_data = output.getvalue()
    output.close()

    filename = f"Timetable_{branch}_Y{year}_S{semester}.csv"

    return Response(
        csv_data,
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename={filename}'
        }
    )
@app.route('/admin/sync_faculties')
@login_required
def admin_sync_faculties():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    added = 0
    professors = User.query.filter_by(role='professor').all()

    for prof in professors:
        fac = Faculty.query.filter_by(email=prof.email).first()

        if not fac:
            new_fac = Faculty(
                id=prof.id,
                name=prof.fullname or prof.username,
                email=prof.email,
                phone=None,
                designation="Professor",
                branches=prof.branch or "CSE,AD"
            )
            db.session.add(new_fac)
            added += 1

    db.session.commit()
    flash(f"Faculty Sync Complete! {added} Professors Added", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/fix_faculties')
@login_required
def fix_faculties():
    if current_user.role != 'admin':
        flash("Access denied", "danger")
        return redirect(url_for("login"))

    professors = User.query.filter_by(role='professor').all()
    added = 0

    for p in professors:
        fac = Faculty.query.filter_by(id=p.id).first()
        if not fac:
            new_fac = Faculty(
                id=p.id,
                name=p.fullname or p.username,
                email=p.email,
                designation="Professor",
                branches=p.branch or "CSE,AD",
                phone=None
            )
            db.session.add(new_fac)
            added += 1

    db.session.commit()
    flash(f"Faculty Sync Complete! {added} record(s) added", "success")
    return redirect(url_for("admin_dashboard"))

# ========== FIXED REPORT GENERATION ROUTES ==========
@app.route('/admin/generate_monthly_attendance', methods=['GET', 'POST'])
@login_required
def generate_monthly_attendance():
    """Generate monthly attendance report in Excel format"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            # Get form data
            branch = request.form.get('branch')
            year = int(request.form.get('year'))
            semester = int(request.form.get('semester'))
            subject_id = int(request.form.get('subject_id'))
            month_year = request.form.get('month_year')

            # Parse month and year
            year_num, month_num = map(int, month_year.split('-'))
            month_name = datetime(year_num, month_num, 1).strftime('%B %Y')

            # Get subject and professor details
            subject = Subject.query.get_or_404(subject_id)
            professor_name = get_professor_for_subject(subject_id)

            # Get working days for the month
            days_in_month = get_working_days_for_month(year_num, month_num)
            day_list = list(days_in_month.keys())

            # Get students for this branch and year
            students = Student.query.filter_by(
                branch=branch,
                year=year
            ).order_by(Student.roll).all()

            # Prepare student data with daily attendance
            students_rows = []
            for idx, student in enumerate(students, 1):
                attendance_by_day = {}

                for day_num, attendance_date in days_in_month.items():
                    attendance = Attendance.query.filter_by(
                        student_id=student.id,
                        subject_id=subject_id,
                        date=attendance_date
                    ).first()

                    attendance_by_day[day_num] = 'P' if (attendance and attendance.status == 'present') else 'A'

                # Calculate totals
                total_present = sum(1 for status in attendance_by_day.values() if status == 'P')
                total_lectures = len(day_list)
                percentage = round((total_present / total_lectures) * 100, 2) if total_lectures > 0 else 0

                students_rows.append({
                    'sr_no': idx,
                    'roll_no': student.roll,
                    'name': student.name,
                    'branch': student.branch,
                    'year': student.year,
                    'attendance_by_day': attendance_by_day,
                    'total_present': total_present,
                    'total_lectures': total_lectures,
                    'percentage': percentage
                })

            # Meta data for Excel
            meta = {
                "college_name": "Shri Balaji Institute of Technology & Management, Betul (M.P.)",
                "report_title": "Monthly Attendance Report",
                "branch": branch,
                "year": year,
                "semester": semester,
                "section": "A",
                "subject_code": subject.code,
                "subject_name": subject.name,
                "faculty_name": professor_name,
                "month_name": month_name
            }

            # Generate filename and path
            filename = f"Monthly_Attendance_{branch}_Y{year}_S{semester}_{subject.code}_{month_year.replace('-', '_')}.xlsx"
            filepath = os.path.join(REPORT_DIR, filename)

            # Generate Excel file
            generate_monthly_attendance_excel(filepath, meta, students_rows, day_list)

            # Save report record
            report = AttendanceReport(
                professor_id=None,
                subject_id=subject_id,
                date=date.today(),
                report_path=filepath,
                report_type='monthly_excel'
            )
            db.session.add(report)
            db.session.commit()

            return send_file(filepath, as_attachment=True, download_name=filename)

        except Exception as e:
            flash(f'Error generating report: {str(e)}', 'danger')
            return redirect(url_for('generate_monthly_attendance'))

    # GET request - show form
    branches = ['CSE', 'AD']
    years = [1, 2, 3, 4]
    subjects = Subject.query.filter_by(is_active=True).order_by(Subject.branch, Subject.semester).all()

    return render_template('admin/monthly_attendance_form.html',
                           branches=branches,
                           years=years,
                           subjects=subjects)


@app.route('/admin/generate_report', methods=['GET', 'POST'])
@login_required
def admin_generate_report():
    if current_user.role != 'admin':
        return redirect(url_for('admin_dashboard'))

    subjects = Subject.query.all()
    today = datetime.now().date()

    if request.method == 'POST':
        try:
            # [OK] MULTIPLE SUBJECTS SUPPORT - Get list of selected subject IDs
            subject_ids = request.form.getlist('subject_ids')

            branch = request.form.get('branch')
            year = int(request.form.get('acad_year'))
            semester = request.form.get('semester', 'all')
            total_working_days = int(request.form.get('total_working_days', 30))
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            last_n = request.form.get('last_n')

            # Date range determination
            if last_n:
                days = int(last_n)
                end_date = today
                start_date = today - timedelta(days=days - 1)
            else:
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                else:
                    start_date = today - timedelta(days=30)  # Default: last 30 days

                if end_date_str:
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                else:
                    end_date = today

            # [OK] DETERMINE SELECTED SUBJECTS
            selected_subjects = []
            report_title = ""

            if 'all' in subject_ids:
                # All subjects in selected scope
                if semester == 'all':
                    # All semesters for the year
                    sem1 = year * 2 - 1
                    sem2 = year * 2
                    selected_subjects = Subject.query.filter(
                        Subject.semester.in_([sem1, sem2])
                    )
                else:
                    # Specific semester
                    selected_subjects = Subject.query.filter_by(semester=int(semester))

                # Filter by branch if selected
                if branch:
                    selected_subjects = selected_subjects.filter(Subject.branch == branch)

                selected_subjects = selected_subjects.order_by(Subject.semester, Subject.code).all()
                report_title = f"All Subjects - Year {year}" + (f" Sem {semester}" if semester != 'all' else "")

            else:
                # Specific selected subjects
                subject_ids_int = [int(sid) for sid in subject_ids]
                selected_subjects = Subject.query.filter(
                    Subject.id.in_(subject_ids_int)
                ).order_by(Subject.semester, Subject.code).all()
                subject_codes = ", ".join([subj.code for subj in selected_subjects[:3]])
                if len(selected_subjects) > 3:
                    subject_codes += f" and {len(selected_subjects) - 3} more"
                report_title = f"Selected Subjects: {subject_codes}"

            if not selected_subjects:
                flash("No subjects found for this selection!", "danger")
                return redirect(url_for('admin_reports'))

            # [OK] Get students based on branch and year
            student_query = Student.query.filter_by(year=year)
            if branch:
                student_query = student_query.filter_by(branch=branch)

            students = student_query.order_by(Student.roll).all()

            if not students:
                flash("No students found for this selection!", "danger")
                return redirect(url_for('admin_reports'))

            # ---------- Excel Work Start ----------
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"

            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            #  total columns = 3 (Sno, Enr, Name) + 3 per subject + 3 (Total LS/LA/%)
            total_cols = 3 + len(selected_subjects) * 3 + 3
            last_col_letter = get_column_letter(total_cols)

            # ---------- HEADER ----------
            ws.merge_cells(f'A1:{last_col_letter}1')
            ws['A1'] = "Shri Balaji Institute of Technology & Management, Betul (M.P.)"
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].font = Font(bold=True, size=14)

            # Department based on branch
            ws.merge_cells(f'A2:{last_col_letter}2')
            dept_name = "Computer Science and Engineering" if branch == "CSE" else "AI & Data Science" if branch == "AD" else "All Departments"
            ws['A2'] = f"Department of {dept_name}"
            ws['A2'].alignment = Alignment(horizontal='center')
            ws['A2'].font = Font(bold=True)

            # Report title with subject info
            ws.merge_cells(f'A3:{last_col_letter}3')
            ws['A3'] = f"Year: {year} | {report_title} | Session: {start_date.year}-{end_date.year}"
            ws['A3'].alignment = Alignment(horizontal='center')
            ws['A3'].font = Font(bold=True, size=12)

            ws.merge_cells(f'A4:{last_col_letter}4')
            ws['A4'] = f"Attendance Record (From {start_date.strftime('%d.%m.%Y')} to {end_date.strftime('%d.%m.%Y')})"
            ws['A4'].alignment = Alignment(horizontal='center')
            ws['A4'].font = Font(bold=True)

            ws.merge_cells(f'A5:{last_col_letter}5')
            ws['A5'] = f"Total Working Days - {total_working_days}"
            ws['A5'].alignment = Alignment(horizontal='center')

            # ---------- TABLE HEADER ROW ----------
            header_row = 7

            headers = ["S. No", "Enrollment No", "Name of Student"]
            for sub in selected_subjects:
                # har subject ke 3 column: LS, LA, %
                headers.extend([
                    f"{sub.code} LS",
                    f"{sub.code} LA",
                    f"{sub.code} %"
                ])

            headers.extend(["Total Lectures Scheduled", "Total Attendance", "Percentage Attendance"])

            ws.append(headers)

            # ---------- DATA ROWS ----------
            for i, stu in enumerate(students, start=1):
                # Safe roll number extraction
                roll_no = (
                        getattr(stu, 'enrollment_no', None)
                        or getattr(stu, 'enrollment', None)
                        or getattr(stu, 'university_roll', None)
                        or getattr(stu, 'roll_no', None)
                        or getattr(stu, 'roll', None)
                        or str(stu.id)
                )

                name = getattr(stu, 'name', None) or getattr(stu, 'fullname', None) or ""

                row = [i, roll_no, name]

                total_scheduled_all = 0
                total_attended_all = 0

                for sub in selected_subjects:
                    scheduled = total_working_days

                    # Count attendance for each subject
                    attended = Attendance.query.filter(
                        Attendance.student_id == stu.id,
                        Attendance.subject_id == sub.id,
                        Attendance.date >= start_date,
                        Attendance.date <= end_date,
                        Attendance.status == 'present'
                    ).count()

                    total_scheduled_all += scheduled
                    total_attended_all += attended

                    percent = round((attended / scheduled) * 100, 2) if scheduled > 0 else 0
                    row.extend([scheduled, attended, percent])

                total_percent = round((total_attended_all / total_scheduled_all) * 100, 2) if total_scheduled_all else 0
                row.extend([total_scheduled_all, total_attended_all, total_percent])

                ws.append(row)

            # ---------- BORDER + ALIGNMENT ----------
            for row in ws.iter_rows(
                    min_row=header_row,
                    max_row=ws.max_row,
                    min_col=1,
                    max_col=total_cols
            ):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # [OK] Column width optimization
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 28
            for col_idx in range(4, total_cols + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12

            # ---------- SAVE FILE ----------
            reports_dir = os.path.join(app.root_path, 'reports')
            os.makedirs(reports_dir, exist_ok=True)

            # Generate filename based on selection
            if 'all' in subject_ids:
                filename = f"ATT_{branch if branch else 'ALL'}_Y{year}_S{semester if semester != 'all' else 'ALL'}_{today.strftime('%Y-%m-%d')}.xlsx"
            else:
                subject_codes = "_".join([subj.code for subj in selected_subjects[:3]])
                if len(selected_subjects) > 3:
                    subject_codes += f"_and_{len(selected_subjects) - 3}_more"
                filename = f"ATT_{branch if branch else 'ALL'}_Y{year}_{subject_codes}_{today.strftime('%Y-%m-%d')}.xlsx"

            filepath = os.path.join(reports_dir, filename)

            wb.save(filepath)

            # DB me path save karo (AttendanceReport model)
            report = AttendanceReport(
                professor_id=None,
                subject_id=None,  # multiple subjects  None
                date=today,
                report_path=filepath,
                report_type='multiple_subjects_excel'
            )
            db.session.add(report)
            db.session.commit()

            flash(f"[OK] Attendance report generated successfully for {len(selected_subjects)} subjects!", "success")
            return send_file(filepath, as_attachment=True, download_name=filename)

        except Exception as e:
            flash(f"Error generating report: {str(e)}", "danger")
            return redirect(url_for('admin_reports'))

    # GET request - show form
    years = [2, 3, 4]  # Academic years
    return render_template("admin/generate_report.html", subjects=subjects, today=today, years=years)


def generate_custom_csv_report():
    """Generate custom CSV report with flexible options"""
    try:
        # Get form data
        branch = request.form.get('branch')
        year = int(request.form.get('year'))
        semester = int(request.form.get('semester'))
        subject_id = request.form.get('subject_id')

        # Date range options
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        last_n = request.form.get('last_n')

        # Validate inputs
        if not branch or not year or not semester:
            flash('Branch, Year and Semester are required', 'danger')
            return redirect(url_for('admin_generate_report'))

        today = date.today()

        # Determine date range
        if last_n:
            # Last N days
            days = int(last_n)
            end_date = today
            start_date = today - timedelta(days=days - 1)
        else:
            # Custom date range
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            else:
                start_date = today - timedelta(days=14)  # Default: last 14 days

            if end_date_str:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                end_date = today

        if start_date > end_date:
            flash('Start date cannot be greater than end date', 'danger')
            return redirect(url_for('admin_generate_report'))

        # Determine which subjects to include
        all_subjects = Subject.query.filter_by(
            branch=branch,
            semester=semester,
            is_active=True
        ).all()

        subject_ids = []
        single_subject = None
        report_scope = ""

        if subject_id != 'all':
            # Single subject report
            single_subject = Subject.query.get(int(subject_id))
            if not single_subject:
                flash('Selected subject not found', 'danger')
                return redirect(url_for('admin_generate_report'))
            subject_ids = [single_subject.id]
            report_scope = f"Single Subject: {single_subject.code}"
        else:
            # All subjects in the semester
            subject_ids = [s.id for s in all_subjects]
            report_scope = f"All Subjects in Semester {semester}"

        if not subject_ids:
            flash('No subjects found for the selected semester', 'warning')
            return redirect(url_for('admin_generate_report'))

        # Build report data
        report_rows = []
        total_students_processed = 0

        for subject_id in subject_ids:
            subject = Subject.query.get(subject_id)
            if not subject:
                continue

            # Get students for this subject's branch and year
            students = Student.query.filter_by(
                branch=subject.branch,
                year=year
            ).order_by(Student.roll).all()

            for student in students:
                # Count attendance for this student+subject in date range
                present_count = Attendance.query.filter(
                    Attendance.student_id == student.id,
                    Attendance.subject_id == subject.id,
                    Attendance.date.between(start_date, end_date),
                    Attendance.status == 'present'
                ).count()

                # Count total classes for this subject in date range
                total_classes = db.session.query(
                    func.count(func.distinct(Attendance.date))
                ).filter(
                    Attendance.subject_id == subject.id,
                    Attendance.date.between(start_date, end_date)
                ).scalar() or 0

                # Calculate percentage
                percentage = round((present_count / total_classes) * 100, 2) if total_classes > 0 else 0.0

                # Determine status
                if percentage >= 75:
                    status_label = 'Good'
                elif percentage >= 50:
                    status_label = 'Average'
                else:
                    status_label = 'Low'

                report_rows.append({
                    'Roll No': student.roll,
                    'Student Name': student.name,
                    'Branch': student.branch,
                    'Year': student.year,
                    'Subject Code': subject.code,
                    'Subject Name': subject.name,
                    'Semester': subject.semester,
                    'Total Present': present_count,
                    'Total Lectures': total_classes,
                    'Attendance %': percentage,
                    'Status': status_label,
                    'Period': f"{start_date} to {end_date}"
                })
                total_students_processed += 1

        if not report_rows:
            flash('No attendance data found for the selected period', 'warning')
            return redirect(url_for('admin_generate_report'))

        # Ensure reports directory exists
        os.makedirs(REPORT_DIR, exist_ok=True)

        # Generate filename
        if single_subject:
            base_name = f"ATT_{branch}_Y{year}_S{semester}_{single_subject.code}"
        else:
            base_name = f"ATT_{branch}_Y{year}_S{semester}_ALL_SUBJECTS"

        filename = f"{base_name}_{today}.csv"
        filepath = os.path.join(REPORT_DIR, filename)

        # Write CSV file
        fieldnames = list(report_rows[0].keys())
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report_rows)

        # Save report record
        new_report = AttendanceReport(
            professor_id=None,
            subject_id=single_subject.id if single_subject else None,
            date=today,
            report_path=filepath,
            report_type='custom_csv'
        )
        db.session.add(new_report)
        db.session.commit()

        flash(f'[OK] CSV report generated successfully! Processed {total_students_processed} student records.',
              'success')
        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        db.session.rollback()
        flash(f'Error generating CSV report: {str(e)}', 'danger')
        return redirect(url_for('admin_generate_report'))


@app.route('/get_subjects_by_year/<int:year>')
@login_required
def get_subjects_by_year(year):
    """API endpoint to get subjects by academic year and branch"""
    try:
        branch = request.args.get('branch', '').strip()

        # Calculate semesters for the year
        sem1 = year * 2 - 1
        sem2 = year * 2

        # Build query
        query = Subject.query.filter(Subject.semester.in_([sem1, sem2]))

        if branch:
            query = query.filter(Subject.branch == branch)

        subjects = query.order_by(Subject.semester, Subject.code).all()

        subjects_data = []
        for subject in subjects:
            subjects_data.append({
                'id': subject.id,
                'code': subject.code,
                'name': subject.name,
                'branch': subject.branch,
                'semester': subject.semester,
                'is_active': subject.is_active
            })

        return jsonify({
            'success': True,
            'subjects': subjects_data,
            'count': len(subjects_data)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/admin/reports')
@login_required
def admin_reports():
    """View all generated reports"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    reports_data = []
    reports = AttendanceReport.query.order_by(AttendanceReport.date.desc()).all()

    for report in reports:
        professor = User.query.get(report.professor_id) if report.professor_id else None
        subject = Subject.query.get(report.subject_id) if report.subject_id else None

        reports_data.append({
            'id': report.id,
            'date': report.date,
            'professor_name': professor.fullname if professor else 'Admin',
            'professor_email': professor.email if professor else 'admin@college.com',
            'subject_name': subject.name if subject else ('All Subjects' if report.subject_id is None else ''),
            'subject_code': subject.code if subject else ('ALL' if report.subject_id is None else ''),
            'filename': os.path.basename(report.report_path) if report.report_path else '',
            'report_path': report.report_path,
            'has_file': bool(report.report_path and os.path.exists(report.report_path)),
            'report_type': getattr(report, 'report_type', 'custom_csv')
        })

    return render_template('admin/reports.html', reports=reports_data)


@app.route('/delete_report/<int:report_id>', methods=['POST'])
@login_required
def delete_report(report_id):
    """Delete a specific report"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    report = AttendanceReport.query.get_or_404(report_id)

    if report.report_path and os.path.exists(report.report_path):
        try:
            os.remove(report.report_path)
        except:
            pass

    db.session.delete(report)
    db.session.commit()

    flash('Report deleted successfully', 'success')
    return redirect(url_for('admin_reports'))


@app.route('/download_report/<int:report_id>')
@login_required
def download_report(report_id):
    """Download a specific report"""
    report = AttendanceReport.query.get_or_404(report_id)

    if current_user.role == 'professor' and report.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_dashboard'))

    if report.report_path and os.path.exists(report.report_path):
        return send_file(report.report_path, as_attachment=True)
    else:
        flash('Report file not found', 'danger')
        return redirect(url_for('admin_reports' if current_user.role == 'admin' else 'prof_reports'))


@app.route('/admin/configure_subjects', methods=['GET', 'POST'])
@login_required
def configure_subjects():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    branches = ['CSE', 'AD']
    years = [1, 2, 3, 4]

    selected_branch = request.values.get('branch', 'CSE')
    selected_year = request.values.get('year', '')
    selected_year_int = None

    if selected_year:
        try:
            selected_year_int = int(selected_year)
        except ValueError:
            selected_year_int = None

    subjects = []
    sems = []

    if selected_branch and selected_year_int:
        sem1 = selected_year_int * 2 - 1
        sem2 = selected_year_int * 2
        sems = [sem1, sem2]

        subjects = Subject.query.filter(
            Subject.branch == selected_branch,
            Subject.semester.in_(sems)
        ).order_by(Subject.semester, Subject.code).all()

        if request.method == 'POST':
            selected_ids = request.form.getlist('subject_ids')
            selected_ids = {int(sid) for sid in selected_ids}

            try:
                for subj in subjects:
                    subj.is_active = subj.id in selected_ids
                db.session.commit()
                flash('Offered subjects updated successfully!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating subjects: {e}', 'danger')

            return redirect(url_for(
                'configure_subjects',
                branch=selected_branch,
                year=selected_year_int
            ))

    return render_template(
        'admin/configure_subjects.html',
        branches=branches,
        years=years,
        selected_branch=selected_branch,
        selected_year=selected_year_int,
        subjects=subjects,
        semesters=sems
    )


@app.route('/admin/allot_subject', methods=['POST'])
@login_required
def admin_allot_subject():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    try:
        prof_id = int(request.form.get('prof_id'))
        subj_id = int(request.form.get('subj_id'))
    except (ValueError, TypeError):
        flash('Invalid professor or subject', 'danger')
        return redirect(url_for('admin_dashboard'))

    existing_allotment = ProfessorSubject.query.filter_by(subject_id=subj_id).first()

    if existing_allotment:
        assigned_prof = User.query.get(existing_allotment.professor_id)
        subject = Subject.query.get(subj_id)
        flash(f'Subject {subject.code} - {subject.name} is already allotted to Professor {assigned_prof.fullname}',
              'warning')
        return redirect(url_for('admin_dashboard'))

    existing_to_same_prof = ProfessorSubject.query.filter_by(professor_id=prof_id, subject_id=subj_id).first()
    if existing_to_same_prof:
        flash('Subject already allotted to this professor', 'warning')
        return redirect(url_for('admin_dashboard'))

    allotment = ProfessorSubject(professor_id=prof_id, subject_id=subj_id)
    db.session.add(allotment)
    db.session.commit()

    professor = User.query.get(prof_id)
    subject = Subject.query.get(subj_id)
    flash(f'Subject {subject.code} - {subject.name} allotted successfully to Professor {professor.fullname}', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/add_subject', methods=['POST'])
@login_required
def add_subject():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    code = request.form.get('code', '').strip()
    name = request.form.get('name', '').strip()
    branch = request.form.get('branch', 'CSE')

    try:
        semester = int(request.form.get('semester', 1))
    except ValueError:
        flash('Invalid semester', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not all([code, name]):
        flash('All fields are required', 'danger')
        return redirect(url_for('admin_dashboard'))

    existing_subject = Subject.query.filter_by(code=code, branch=branch, semester=semester).first()
    if existing_subject:
        flash('Subject with this code already exists for this branch and semester', 'warning')
        return redirect(url_for('admin_dashboard'))

    subject = Subject(code=code, name=name, branch=branch, semester=semester, is_active=True)
    db.session.add(subject)
    db.session.commit()

    flash('Subject added successfully', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/add_professor', methods=['POST'])
@login_required
def add_professor():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    username = request.form.get('username', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    branch = request.form.get('branch', 'CSE')

    if not all([username, name, email, password]):
        flash('All fields are required', 'danger')
        return redirect(url_for('admin_dashboard'))

    if User.query.filter((User.username == username) | (User.email == email)).first():
        flash('Username or email already exists', 'danger')
        return redirect(url_for('admin_dashboard'))

    professor = User(
        username=username,
        fullname=name,
        email=email,
        role='professor',
        branch=branch,
        email_verified=True
    )
    professor.set_password(password)
    db.session.add(professor)
    db.session.commit()

    flash(f'Professor {name} added successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_student', methods=['POST'])
@login_required
def admin_add_student():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    roll = request.form.get('roll', '').strip()
    name = request.form.get('name', '').strip()
    branch = request.form.get('branch', 'CSE')
    try:
        year = int(request.form.get('year', 1))
    except ValueError:
        flash('Invalid year selected', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not all([roll, name, branch, year]):
        flash('All fields are required', 'danger')
        return redirect(url_for('admin_dashboard'))

    # Check if student already exists
    if Student.query.filter_by(roll=roll).first():
        flash(f'Student with Roll Number {roll} already exists', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        # 1. Create Student record
        new_student = Student(
            roll=roll,
            name=name,
            branch=branch,
            year=year
        )
        db.session.add(new_student)

        # 2. Create User record for login
        # Check if user already exists (unlikely if student doesn't, but for safety)
        existing_user = User.query.filter_by(username=roll).first()
        if not existing_user:
            student_user = User(
                username=roll.upper(),
                fullname=name,
                email=f"{roll.lower()}@college.com",
                role='student',
                branch=branch,
                student_roll=roll.upper(),
                email_verified=True
            )
            student_user.set_password(roll.upper()) # Default password is roll number
            db.session.add(student_user)
        
        db.session.commit()
        flash(f'Student {name} added successfully! They can now login with Roll Number as password.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding student: {str(e)}', 'danger')
        print(f"[ERROR] Add Student: {e}")

    return redirect(url_for('admin_dashboard'))


from sqlalchemy import or_   # 👈 Top par (baaki imports ke saath) ek baar add kar dena

from sqlalchemy import or_

@app.route('/admin/delete_professor/<int:prof_id>', methods=['POST'])
@login_required
def delete_professor(prof_id):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    professor = User.query.get_or_404(prof_id)

    if professor.role != 'professor':
        flash('Invalid professor', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        # Assign tests to admin instead of NULL
        admin_user = User.query.filter_by(role='admin').first()
        Test.query.filter_by(professor_id=prof_id).update(
            {Test.professor_id: admin_user.id}, synchronize_session=False
        )

        # Remove professor-subject links
        ProfessorSubject.query.filter_by(professor_id=prof_id).delete()

        # Remove any unused OTP linked to this user
        PasswordResetOTP.query.filter_by(user_id=prof_id).delete(synchronize_session=False)

        # Delete professor user
        db.session.delete(professor)
        db.session.commit()

        flash('Professor deleted successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting professor: {e}', 'danger')
        print(f"[ERROR] Delete Professor: {e}")

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/remove_allotment/<int:allot_id>', methods=['POST'])
@login_required
def remove_allotment(allot_id):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    allotment = ProfessorSubject.query.get_or_404(allot_id)
    db.session.delete(allotment)
    db.session.commit()

    flash('Allotment removed successfully', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/semester', methods=['GET', 'POST'])
@login_required
def admin_semester_management():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    active_semesters = get_all_active_semesters()

    if request.method == 'POST':
        try:
            branch = request.form.get('branch')
            year = int(request.form.get('year'))
            semester_type = request.form.get('semester_type')
            academic_year = int(request.form.get('academic_year'))

            CurrentSemester.query.filter_by(branch=branch, year=year, is_active=True).update({'is_active': False})

            new_semester = CurrentSemester(
                branch=branch,
                year=year,
                semester_type=semester_type,
                academic_year=academic_year,
                is_active=True
            )
            db.session.add(new_semester)
            db.session.commit()

            flash(f'Semester updated for {branch} Year {year}: {academic_year} {semester_type.capitalize()} Semester',
                  'success')
            return redirect(url_for('admin_semester_management'))

        except Exception as e:
            flash(f'Error updating semester: {str(e)}', 'danger')
            return redirect(url_for('admin_semester_management'))

    return render_template('admin/semester_management.html', active_semesters=active_semesters)


@app.route('/admin/fix_student_accounts')
@login_required
def fix_student_accounts():
    if current_user.role != 'admin':
        return "Access denied"

    students = Student.query.all()
    created_count = 0
    fixed_count = 0
    error_count = 0

    result = "<h3>Synchronizing Student User Accounts (Case Normalization)</h3>"

    for student in students:
        try:
            # Normalize student roll in the Student table first
            if student.roll and student.roll != student.roll.upper():
                student.roll = student.roll.upper()
                fixed_count += 1

            # Find user by student_roll (case-insensitive check)
            existing_user = User.query.filter(db.func.lower(User.student_roll) == student.roll.lower()).first()
            if not existing_user:
                # Also check by username as backup
                existing_user = User.query.filter(db.func.lower(User.username) == student.roll.lower()).first()

            if not existing_user:
                student_user = User(
                    username=student.roll.upper(),
                    fullname=student.name,
                    email=f"{student.roll.lower()}@college.com",
                    role='student',
                    branch=student.branch,
                    student_roll=student.roll.upper(),
                    email_verified=True
                )
                student_user.set_password(student.roll.upper())
                db.session.add(student_user)
                created_count += 1
                result += f"<p>[NEW] Created: {student.roll} - {student.name}</p>"
            else:
                # Fix existing user casing and password
                needs_update = False
                if existing_user.username != student.roll.upper():
                    existing_user.username = student.roll.upper()
                    needs_update = True
                if existing_user.student_roll != student.roll.upper():
                    existing_user.student_roll = student.roll.upper()
                    needs_update = True
                
                # Always reset password to uppercase roll number if it's a student to be safe
                existing_user.set_password(student.roll.upper())
                needs_update = True

                if needs_update:
                    fixed_count += 1
                    result += f"<p>[FIX] Updated casing/password: {student.roll}</p>"
                else:
                    result += f"<p>[OK] Already correct: {student.roll}</p>"

        except Exception as e:
            error_count += 1
            result += f"<p style='color: red;'>[ERROR] Error with {student.roll}: {str(e)}</p>"

    if created_count > 0 or fixed_count > 0:
        db.session.commit()
        result += f"<h4 style='color: green;'>[OK] SUCCESS: Created {created_count} and Updated {fixed_count} student accounts!</h4>"
    else:
        result += "<h4> No changes needed (all records already matched)</h4>"

    if error_count > 0:
        result += f"<h4 style='color: red;'>[ERROR] ERRORS: {error_count} records failed</h4>"

    result += f"""
    <hr>
    <h4>Login Verification (Testing):</h4>
    <p>Please try logging in with:</p>
    <ul>
        <li><strong>Username:</strong> {students[0].roll if students else "Any Roll Number"}</li>
        <li><strong>Password:</strong> {students[0].roll if students else "Same Roll Number"}</li>
    </ul>
    <a href="/admin" class="btn btn-secondary" style="padding: 10px 20px; background: #6c757d; color: white; border-radius: 5px; text-decoration: none;">Back to Dashboard</a>
    """

    return result

@app.route('/admin/notices')
@login_required
def admin_notices():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    notices = Notice.query.order_by(Notice.created_at.desc()).all()

    notices_with_creators = []
    for notice in notices:
        creator = User.query.get(notice.created_by)
        notices_with_creators.append({
            'notice': notice,
            'creator_name': creator.fullname if creator else 'Admin'
        })

    from datetime import datetime
    return render_template('admin/notices.html',
                           notices_with_creators=notices_with_creators,
                           datetime=datetime)


@app.route('/admin/create_notice', methods=['POST'])
@login_required
def create_notice():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    try:
        title = request.form.get('title', '').strip()
        message = request.form.get('message', '').strip()
        target_audience = request.form.get('target_audience', 'all')
        branch = request.form.get('branch')
        year = request.form.get('year')
        is_important = 'is_important' in request.form

        if not all([title, message]):
            flash('Title and message are required', 'danger')
            return redirect(url_for('admin_notices'))

        year_int = int(year) if year and year.isdigit() else None

        notice = Notice(
            title=title,
            message=message,
            created_by=current_user.id,
            target_audience=target_audience,
            branch=branch,
            year=year_int,
            is_important=is_important,
            expires_at=datetime.now() + timedelta(days=30)
        )
        db.session.add(notice)
        db.session.commit()

        flash('Notice created successfully!', 'success')
        return redirect(url_for('admin_notices'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error creating notice: {str(e)}', 'danger')
        return redirect(url_for('admin_notices'))


@app.route('/admin/delete_notice/<int:notice_id>', methods=['POST'])
@login_required
def delete_notice(notice_id):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    notice = Notice.query.get_or_404(notice_id)

    try:
        db.session.delete(notice)
        db.session.commit()
        flash('Notice deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting notice: {str(e)}', 'danger')

    return redirect(url_for('admin_notices'))


@app.route('/admin/tests')
@login_required
def admin_tests():
    """Admin test management - view all tests"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    tests = Test.query.order_by(Test.created_at.desc()).all()
    return render_template('admin/tests.html', tests=tests)
from sqlalchemy.exc import SQLAlchemyError, IntegrityError

@app.route('/admin/test/<int:test_id>/delete', methods=['POST'])
@login_required
def admin_delete_test(test_id):
    """Admin delete test with full cascade"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)

    try:
        # ORM cascades handle:
        # - Test.attempts -> TestAttempt
        # - TestAttempt.answers -> StudentAnswer
        # - Test.questions -> Question
        # - Question.student_answers -> StudentAnswer
        # - Test.sections -> QuestionSection

        db.session.delete(test)
        db.session.commit()

        flash('Test and all related data deleted successfully.', 'success')

    except IntegrityError as e:
        db.session.rollback()
        print("IntegrityError while deleting test:", e)
        flash('Database constraint error while deleting test. Please contact developer.', 'danger')

    except SQLAlchemyError as e:
        db.session.rollback()
        print("SQLAlchemyError while deleting test:", e)
        flash(f'Error deleting test: {str(e)}', 'danger')

    return redirect(url_for('admin_tests'))

@app.route('/admin/test/<int:test_id>/results')
@login_required
def admin_test_results(test_id):
    """View test results"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    attempts = TestAttempt.query.filter_by(test_id=test_id, submitted=True).all()

    return render_template('admin/test_results.html',
                           test=test,
                           attempts=attempts)


# ========== PROFESSOR ROUTES ==========
@app.route('/prof')
@login_required
def prof_dashboard():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subjects = []

    for allotment in allotments:
        subject = Subject.query.get(allotment.subject_id)
        if subject:
            subjects.append(subject)

    # Get recent notices for professor
    from datetime import datetime
    recent_notices = Notice.query.filter(
        Notice.is_active == True,
        Notice.expires_at >= datetime.now(),
        db.or_(
            Notice.target_audience == 'all',
            Notice.target_audience == 'professors'
        )
    ).order_by(Notice.created_at.desc()).limit(5).all()

    return render_template('prof/dashboard.html',
                           subjects=subjects,
                           recent_notices=recent_notices,
                           current_user=current_user)


@app.route('/prof/attendance/<int:subject_id>', methods=['GET', 'POST'])
@login_required
def take_attendance(subject_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    allotment = ProfessorSubject.query.filter_by(professor_id=current_user.id, subject_id=subject_id).first()

    if not allotment:
        flash('You are not allotted this subject', 'danger')
        return redirect(url_for('prof_dashboard'))

    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    if request.method == 'POST':
        attendance_date_str = request.form.get('date')
        if not attendance_date_str:
            flash('Please select a date', 'danger')
            return redirect(url_for('take_attendance', subject_id=subject_id))

        attendance_date = date.fromisoformat(attendance_date_str)

        # Delete existing attendance for that date
        Attendance.query.filter_by(subject_id=subject_id, date=attendance_date).delete()

        for student in students:
            status = request.form.get(f'status_{student.id}', 'absent')
            attendance = Attendance(
                student_id=student.id,
                subject_id=subject_id,
                date=attendance_date,
                status=status
            )
            db.session.add(attendance)

        db.session.commit()
        flash('Attendance submitted successfully', 'success')
        return redirect(url_for('prof_dashboard'))

    return render_template('prof/attendance.html',
                           subject=subject,
                           students=students,
                           today=date.today())


@app.route('/prof/bulk_attendance/<int:subject_id>', methods=['GET', 'POST'])
@login_required
def bulk_attendance(subject_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    if request.method == 'POST':
        attendance_date_str = request.form.get('date')
        if not attendance_date_str:
            flash('Please select a date', 'danger')
            return redirect(url_for('bulk_attendance', subject_id=subject_id))

        attendance_date = date.fromisoformat(attendance_date_str)
        bulk_action = request.form.get('bulk_action', 'present')

        # Delete existing attendance for that date
        Attendance.query.filter_by(subject_id=subject_id, date=attendance_date).delete()

        for student in students:
            attendance = Attendance(
                student_id=student.id,
                subject_id=subject_id,
                date=attendance_date,
                status=bulk_action
            )
            db.session.add(attendance)

        db.session.commit()
        flash(f'Bulk attendance marked as {bulk_action} for all students', 'success')
        return redirect(url_for('prof_dashboard'))

    return render_template('prof/bulk_attendance.html',
                           subject=subject,
                           students=students,
                           today=date.today())


@app.route('/prof/mid_term_marks/<int:subject_id>', methods=['GET', 'POST'])
@login_required
def enter_mid_term_marks(subject_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    allotment = ProfessorSubject.query.filter_by(professor_id=current_user.id, subject_id=subject_id).first()

    if not allotment:
        flash('You are not allotted this subject', 'danger')
        return redirect(url_for('prof_dashboard'))

    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    current_semester = get_active_semester_for_branch_year(subject.branch, academic_year)

    if request.method == 'POST':
        try:
            total_marks = float(request.form.get('total_marks', 30))
            exam_type = request.form.get('exam_type', 'mid_term')

            marks_entered = 0
            errors = []

            for student in students:
                marks_key = f'marks_{student.id}'
                if marks_key in request.form and request.form[marks_key].strip():
                    marks_obtained = float(request.form[marks_key])

                    if marks_obtained > total_marks:
                        errors.append(
                            f"Roll No {student.roll}: Obtained marks ({marks_obtained}) cannot exceed total marks ({total_marks})")
                        continue

                    existing_marks = MidTermMarks.query.filter_by(
                        student_id=student.id,
                        subject_id=subject_id,
                        exam_type=exam_type,
                        semester=subject.semester
                    ).first()

                    if existing_marks:
                        existing_marks.marks_obtained = marks_obtained
                        existing_marks.total_marks = total_marks
                        existing_marks.updated_at = datetime.now()

                    else:
                        mid_term_marks = MidTermMarks(
                            student_id=student.id,
                            subject_id=subject_id,
                            professor_id=current_user.id,
                            marks_obtained=marks_obtained,
                            total_marks=total_marks,
                            exam_type=exam_type,
                            semester=subject.semester,
                            academic_year=current_semester.academic_year if current_semester else datetime.now().year
                        )
                        db.session.add(mid_term_marks)

                    marks_entered += 1

            if errors:
                for error in errors:
                    flash(error, 'danger')
            else:
                db.session.commit()
                flash(f'[OK] Mid-term marks entered for {marks_entered} students! Total marks: {total_marks}',
                      'success')
                return redirect(url_for('prof_dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error entering marks: {str(e)}', 'danger')

    existing_marks = {}
    marks_entries = MidTermMarks.query.filter_by(subject_id=subject_id, exam_type='mid_term').all()

    for entry in marks_entries:
        existing_marks[entry.student_id] = {
            'marks_obtained': entry.marks_obtained,
            'total_marks': entry.total_marks
        }

    return render_template('prof/mid_term_marks.html',
                           subject=subject,
                           students=students,
                           existing_marks=existing_marks,
                           academic_year=academic_year)


@app.route('/prof/view_marks/<int:subject_id>')
@login_required
def view_subject_marks(subject_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    allotment = ProfessorSubject.query.filter_by(professor_id=current_user.id, subject_id=subject_id).first()

    if not allotment:
        flash('You are not allotted this subject', 'danger')
        return redirect(url_for('prof_dashboard'))

    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    marks_data = MidTermMarks.query.filter_by(subject_id=subject_id, exam_type='mid_term').all()
    marks_map = {mark.student_id: mark for mark in marks_data}

    return render_template('prof/view_marks.html',
                           subject=subject,
                           students=students,
                           marks_map=marks_map,
                           academic_year=academic_year)


@app.route('/prof/reports')
@login_required
def prof_reports():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subj_map = {}

    for allotment in allotments:
        subject = Subject.query.get(allotment.subject_id)
        if subject:
            reports = AttendanceReport.query.filter_by(
                subject_id=subject.id,
                professor_id=current_user.id
            ).order_by(AttendanceReport.date.desc()).all()

            subj_map[subject.id] = {
                'subject': subject,
                'dates': [{
                    'date': r.date,
                    'date_str': r.date.strftime('%d-%m-%Y'),
                    'date_iso': r.date.isoformat(),
                    'report_id': r.id,
                    'file_exists': bool(r.report_path and os.path.exists(r.report_path)),
                    'path': r.report_path
                } for r in reports]
            }

    return render_template('prof/reports.html', subj_map=subj_map)


@app.route('/prof/report/<int:subject_id>')
@login_required
def generate_subject_report(subject_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    days = int(request.args.get('days', 15))

    if request.args.get('download') == 'true':
        return download_subject_report(subject_id)

    start_date = date.today() - timedelta(days=days - 1)
    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    reports = []
    for student in students:
        present_count = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.subject_id == subject_id,
            Attendance.date.between(start_date, date.today()),
            Attendance.status == 'present'
        ).count()

        percentage = round((present_count / days) * 100, 2) if days > 0 else 0

        reports.append({
            'roll': student.roll,
            'name': student.name,
            'present': present_count,
            'percent': percentage
        })

    return render_template('prof/report.html',
                           subject=subject,
                           reports=reports,
                           days=days)


def download_subject_report(subject_id):
    """Download subject report as CSV file"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    subject = Subject.query.get_or_404(subject_id)
    days = int(request.args.get('days', 15))
    start_date = date.today() - timedelta(days=days - 1)

    academic_year = (subject.semester + 1) // 2
    students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

    report_data = []
    for student in students:
        present_count = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.subject_id == subject_id,
            Attendance.date.between(start_date, date.today()),
            Attendance.status == 'present'
        ).count()

        percentage = round((present_count / days) * 100, 2) if days > 0 else 0

        report_data.append({
            'roll_no': student.roll,
            'name': student.name,
            'present_days': present_count,
            'total_days': days,
            'attendance_percentage': percentage,
            'status': 'Good' if percentage >= 75 else 'Average' if percentage >= 50 else 'Low'
        })

    import csv
    import os

    filename = f"Attendance_Report_{subject.code}_{date.today()}_{days}days.csv"
    filepath = os.path.join(REPORT_DIR, filename)

    try:
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['roll_no', 'name', 'present_days', 'total_days', 'attendance_percentage', 'status']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report_data)

        report = AttendanceReport(
            professor_id=current_user.id,
            subject_id=subject_id,
            date=date.today(),
            report_path=filepath
        )
        db.session.add(report)
        db.session.commit()

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'danger')
        return redirect(url_for('generate_subject_report', subject_id=subject_id))


@app.route('/prof/all_subjects_report')
@login_required
def prof_all_subjects_report():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subjects = [Subject.query.get(a.subject_id) for a in allotments if Subject.query.get(a.subject_id)]

    if not subjects:
        flash('No subjects allotted', 'warning')
        return redirect(url_for('prof_dashboard'))

    days = int(request.args.get('days', 30))
    end_date = date.today()
    start_date = end_date - timedelta(days=days - 1)

    report_data = []

    for subject in subjects:
        academic_year = (subject.semester + 1) // 2
        students = Student.query.filter_by(branch=subject.branch, year=academic_year).order_by(Student.roll).all()

        for student in students:
            present_count = Attendance.query.filter(
                Attendance.student_id == student.id,
                Attendance.subject_id == subject.id,
                Attendance.date.between(start_date, end_date),
                Attendance.status == 'present'
            ).count()

            total_classes = db.session.query(db.func.count(db.func.distinct(Attendance.date))).filter(
                Attendance.subject_id == subject.id,
                Attendance.date.between(start_date, end_date)
            ).scalar() or 0

            percentage = round((present_count / total_classes) * 100, 2) if total_classes > 0 else 0

            report_data.append({
                'roll_no': student.roll,
                'name': student.name,
                'branch': student.branch,
                'year': student.year,
                'subject_code': subject.code,
                'subject_name': subject.name,
                'semester': subject.semester,
                'present_days': present_count,
                'total_days': total_classes,
                'attendance_percentage': percentage,
                'status': 'Good' if percentage >= 75 else 'Average' if percentage >= 50 else 'Low',
                'period': f"{start_date} to {end_date}"
            })

    import csv
    import os

    filename = f"Professor_{current_user.username}_All_Subjects_Report_{date.today()}.csv"
    filepath = os.path.join(REPORT_DIR, filename)

    if report_data:
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = report_data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report_data)

        report = AttendanceReport(
            professor_id=current_user.id,
            subject_id=None,
            date=date.today(),
            report_path=filepath
        )
        db.session.add(report)
        db.session.commit()

        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        flash('No attendance data found for the selected period', 'warning')
        return redirect(url_for('prof_dashboard'))


@app.route('/prof/notes')
@login_required
def prof_notes():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subjects = [Subject.query.get(a.subject_id) for a in allotments if Subject.query.get(a.subject_id)]

    notes = Notes.query.filter_by(professor_id=current_user.id).order_by(Notes.uploaded_at.desc()).all()

    return render_template('prof/notes.html',
                           subjects=subjects,
                           notes=notes,
                           current_user=current_user)


@app.route('/prof/upload_note', methods=['POST'])
@login_required
def upload_note():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    try:
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        subject_id = request.form.get('subject_id')
        file = request.files.get('file')

        if not all([title, subject_id, file]):
            flash('Title, subject and file are required', 'danger')
            return redirect(url_for('prof_notes'))

        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('prof_notes'))

        if not allowed_file(file.filename):
            flash('File type not allowed. Allowed types: PDF, DOC, DOCX, PPT, PPTX, TXT, JPG, PNG', 'danger')
            return redirect(url_for('prof_notes'))

        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        if file_size > MAX_FILE_SIZE:
            flash('File size too large. Maximum 16MB allowed.', 'danger')
            return redirect(url_for('prof_notes'))

        filename = secure_filename(file.filename)
        unique_filename = f"{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
        file_path = os.path.join(NOTES_FOLDER, unique_filename)
        file.save(file_path)

        note = Notes(
            professor_id=current_user.id,
            subject_id=subject_id,
            title=title,
            description=description,
            file_path=file_path,
            file_name=filename,
            file_size=file_size
        )
        db.session.add(note)
        db.session.commit()

        flash('Note uploaded successfully! Students can now download it.', 'success')
        return redirect(url_for('prof_notes'))

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] ERROR in upload_note: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error uploading note: {str(e)}', 'danger')
        return redirect(url_for('prof_notes'))


@app.route('/prof/delete_note/<int:note_id>', methods=['POST'])
@login_required
def delete_note(note_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    note = Notes.query.get_or_404(note_id)

    if note.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_notes'))

    try:
        if note.file_path and os.path.exists(note.file_path):
            os.remove(note.file_path)

        db.session.delete(note)
        db.session.commit()
        flash('Note deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting note: {str(e)}', 'danger')

    return redirect(url_for('prof_notes'))


@app.route('/prof/notices')
@login_required
def prof_notices():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    from datetime import datetime
    notices = Notice.query.filter(
        Notice.is_active == True,
        Notice.expires_at >= datetime.now(),
        db.or_(
            Notice.target_audience == 'all',
            Notice.target_audience == 'professors'
        )
    ).order_by(Notice.created_at.desc()).all()

    for notice in notices:
        mark_notice_as_seen(notice.id, current_user.id)

    return render_template('prof/notices.html', notices=notices)


# ========== SMART TEST SYSTEM ROUTES ==========
@app.route('/prof/tests', methods=['GET'])
@login_required
def prof_tests():
    """Professor test management - COMBINED PAGE"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    # Get professor's subjects
    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subjects = [Subject.query.get(a.subject_id) for a in allotments if Subject.query.get(a.subject_id)]

    # Get professor's tests
    tests = Test.query.filter_by(professor_id=current_user.id).order_by(Test.created_at.desc()).all()

    return render_template('prof/tests_combined.html',
                           tests=tests,
                           subjects=subjects,
                           current_user=current_user,
                           datetime=datetime)


@app.route('/prof/test/<int:test_id>/questions')
@login_required
def manage_test_questions(test_id):
    """Manage questions for a specific test - COMBINED VIEW"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    questions = Question.query.filter_by(test_id=test_id).order_by(Question.question_order).all()

    # Get all tests for the professor (for the combined view)
    all_tests = Test.query.filter_by(professor_id=current_user.id).order_by(Test.created_at.desc()).all()
    allotments = ProfessorSubject.query.filter_by(professor_id=current_user.id).all()
    subjects = [Subject.query.get(a.subject_id) for a in allotments if Subject.query.get(a.subject_id)]

    return render_template('prof/tests_combined.html',
                           test=test,
                           questions=questions,
                           tests=all_tests,
                           subjects=subjects,
                           current_user=current_user,
                           datetime=datetime)


@app.route('/prof/create_test', methods=['POST'])
@login_required
def create_test():
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    try:
        # Get form data
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        subject_id = request.form.get('subject_id')
        duration = int(request.form.get('duration_minutes', 60))
        total_marks = int(request.form.get('total_marks', 100))
        start_time_str = request.form.get('start_time')
        instructions = request.form.get('instructions', '').strip()

        # Security code fields
        require_security_code = 'require_security_code' in request.form
        security_code = request.form.get('security_code') if require_security_code else None

        # Test settings
        auto_submit = 'auto_submit' in request.form
        prevent_tab_switch = 'prevent_tab_switch' in request.form
        allow_retake = 'allow_retake' in request.form

        # Validation
        if not all([title, subject_id, start_time_str]):
            flash('All required fields must be filled', 'danger')
            return redirect(url_for('prof_tests'))

        # FIXED: TIMEZONE HANDLING - Professor enters IST time, store as UTC
        try:
            # Parse the datetime string (assuming format: YYYY-MM-DDTHH:MM)
            start_time_naive = datetime.fromisoformat(start_time_str)

            # Assume professor is entering IST time, convert to UTC for storage
            start_time_ist = IST.localize(start_time_naive)
            start_time_utc = convert_to_utc(start_time_ist)

            # Calculate end time in UTC
            end_time_utc = start_time_utc + timedelta(minutes=duration)

            print(f"⏰ TIME CONVERSION DEBUG:")
            print(f"  Professor entered: {start_time_naive} (assumed IST)")
            print(f"  As IST: {start_time_ist}")
            print(f"  Stored as UTC: {start_time_utc}")
            print(f"  End time UTC: {end_time_utc}")
            print(f"  Duration: {duration} minutes")

        except ValueError as e:
            flash(f'Invalid date format: {str(e)}. Use format: YYYY-MM-DDTHH:MM', 'danger')
            return redirect(url_for('prof_tests'))

        # Check if security code is unique
        if require_security_code and security_code:
            existing_test = Test.query.filter_by(security_code=security_code).first()
            if existing_test:
                flash('Security code already exists. Please generate a new one.', 'danger')
                return redirect(url_for('prof_tests'))

        # Create test with UTC times
        test = Test(
            title=title,
            description=description,
            subject_id=subject_id,
            professor_id=current_user.id,
            total_marks=total_marks,
            duration_minutes=duration,

            # FIXED: Store all times in UTC
            start_time=start_time_utc,
            end_time=end_time_utc,
            available_from=start_time_utc,
            available_until=end_time_utc,

            # Security fields
            security_code=security_code,
            require_security_code=require_security_code,
            security_code_verified=False,

            # Settings
            auto_submit=auto_submit,
            prevent_tab_switch=prevent_tab_switch,
            allow_retake=allow_retake,
            is_active=True,
            instructions=instructions,
            question_types='mcq',
            test_type='mixed',
            status='draft'
        )

        db.session.add(test)
        db.session.commit()

        # Success message with timing info
        start_time_display = convert_to_ist(start_time_utc).strftime('%d %b %Y, %I:%M %p IST')
        end_time_display = convert_to_ist(end_time_utc).strftime('%d %b %Y, %I:%M %p IST')

        flash_msg = f'[OK] Test created successfully! Timing: {start_time_display} to {end_time_display}'

        if require_security_code and security_code:
            flash_msg += f' | Security Code: {security_code}'

        flash(flash_msg, 'success')

        return redirect(url_for('manage_test_questions', test_id=test.id))

    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR in create_test: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error creating test: {str(e)}', 'danger')
        return redirect(url_for('prof_tests'))
@app.route('/fix_all_test_times')
def fix_all_test_times():
    """Convert all existing test times from IST to UTC"""
    tests = Test.query.all()
    fixed_count = 0

    for test in tests:
        if test.available_from and test.available_until:
            # Assume existing times are in IST, convert to UTC
            available_from_ist = IST.localize(test.available_from)
            available_until_ist = IST.localize(test.available_until)

            test.available_from = convert_to_utc(available_from_ist)
            test.available_until = convert_to_utc(available_until_ist)
            fixed_count += 1

    db.session.commit()
    return f"Fixed {fixed_count} test times (IST to UTC conversion)"
@app.route('/prof/test/<int:test_id>/add_question', methods=['POST'])
@login_required
def add_question(test_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    try:
        question_text = request.form.get('question_text', '').strip()
        option_a = request.form.get('option_a', '').strip()
        option_b = request.form.get('option_b', '').strip()
        option_c = request.form.get('option_c', '').strip()
        option_d = request.form.get('option_d', '').strip()
        correct_answer = request.form.get('correct_answer', '').strip().upper()
        marks = int(request.form.get('marks', 1))
        question_order = int(request.form.get('question_order', 0))

        if not all([question_text, option_a, option_b, option_c, option_d, correct_answer]):
            flash('All question fields are required', 'danger')
            return redirect(url_for('manage_test_questions', test_id=test_id))

        if correct_answer not in ['A', 'B', 'C', 'D']:
            flash('Correct answer must be A, B, C, or D', 'danger')
            return redirect(url_for('manage_test_questions', test_id=test_id))

        question = Question(
            test_id=test_id,
            question_text=question_text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            correct_answer=correct_answer,
            marks=marks,
            question_order=question_order
        )
        db.session.add(question)
        db.session.commit()

        flash('Question added successfully!', 'success')
        return redirect(url_for('manage_test_questions', test_id=test_id))

    except Exception as e:
        db.session.rollback()
        flash(f'Error adding question: {str(e)}', 'danger')
        return redirect(url_for('manage_test_questions', test_id=test_id))


@app.route('/prof/test/<int:test_id>/preview')
@login_required
def preview_test(test_id):
    """Preview complete test"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    questions = Question.query.filter_by(test_id=test_id).order_by(Question.question_order).all()
    total_questions = len(questions)
    total_marks = sum(q.marks for q in questions)

    return render_template('prof/preview_test.html',
                           test=test,
                           questions=questions,
                           total_questions=total_questions,
                           total_marks=total_marks)


@app.route('/prof/test/<int:test_id>/toggle_status', methods=['POST'])
@login_required
def toggle_test_status(test_id):
    """Activate/deactivate test"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    try:
        test.is_active = not test.is_active
        db.session.commit()

        status = "activated" if test.is_active else "deactivated"
        flash(f'Test {status} successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error updating test status: {str(e)}', 'danger')

    return redirect(url_for('manage_test_questions', test_id=test_id))


@app.route('/prof/test/<int:test_id>/results')
@login_required
def view_test_results(test_id):
    """View test results for professor"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    attempts = TestAttempt.query.filter_by(test_id=test_id, submitted=True).all()
    return render_template('prof/test_results.html',
                           test=test,
                           attempts=attempts)


@app.route('/prof/test/<int:test_id>/delete_question/<int:question_id>', methods=['POST'])
@login_required
def delete_question(test_id, question_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    question = Question.query.get_or_404(question_id)
    if question.test_id != test_id:
        flash('Invalid question', 'danger')
        return redirect(url_for('manage_test_questions', test_id=test_id))

    try:
        # Delete associated student answers first
        StudentAnswer.query.filter_by(question_id=question_id).delete()

        db.session.delete(question)
        db.session.commit()
        flash('Question deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting question: {str(e)}', 'danger')

    return redirect(url_for('manage_test_questions', test_id=test_id))


@app.route('/prof/test/<int:test_id>/delete', methods=['POST'])
@login_required
def delete_test(test_id):
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    try:
        # Delete associated records
        TestAttempt.query.filter_by(test_id=test_id).delete()
        Question.query.filter_by(test_id=test_id).delete()
        db.session.delete(test)
        db.session.commit()
        flash('Test deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting test: {str(e)}', 'danger')

    return redirect(url_for('prof_tests'))


@app.route('/prof/test/<int:test_id>/publish', methods=['POST'])
@login_required
def publish_test(test_id):
    """Publish test for students - FIXED VERSION"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    if not test.questions:
        flash('Cannot publish test without questions', 'danger')
        return redirect(url_for('manage_test_questions', test_id=test_id))

    try:
        # [OK] CRITICAL FIX: Set all required fields for student visibility
        test.status = 'published'
        test.is_active = True

        # [OK] Ensure timing is set correctly
        if not test.available_from:
            test.available_from = test.start_time
        if not test.available_until:
            test.available_until = test.end_time

        # [OK] Update total marks based on actual questions
        test.total_marks = test.calculated_total_marks

        db.session.commit()

        flash('[OK] Test published successfully! Students can now see it.', 'success')

        # [OK] Debug information
        print(f"[SUCCESS] Test Published: {test.title}")
        print(f"[SUCCESS] Status: {test.status}")
        print(f"[SUCCESS] Available From: {test.available_from}")
        print(f"[SUCCESS] Available Until: {test.available_until}")
        print(f"[SUCCESS] Is Active: {test.is_active}")
        print(f"[SUCCESS] Total Marks: {test.total_marks}")
        print(f"[SUCCESS] Questions Count: {len(test.questions)}")

    except Exception as e:
        db.session.rollback()
        flash(f'Error publishing test: {str(e)}', 'danger')

    return redirect(url_for('manage_test_questions', test_id=test_id))


@app.route('/prof/test/<int:test_id>/unpublish', methods=['POST'])
@login_required
def unpublish_test(test_id):
    """Unpublish test"""
    if current_user.role != 'professor':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    test = Test.query.get_or_404(test_id)
    if test.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_tests'))

    try:
        test.status = 'draft'
        db.session.commit()
        flash('Test unpublished successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error unpublishing test: {str(e)}', 'danger')

    return redirect(url_for('manage_test_questions', test_id=test_id))


# ========== STUDENT ROUTES ==========
@app.route('/student')
@login_required
def student_dashboard():
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash('Student record not found', 'danger')
        return redirect(url_for('logout'))

    # Get current semesters
    current_semesters = get_student_current_semesters(student.branch, student.year)

    # [OK] FIX: current_semester_config define karo
    current_semester_config = get_active_semester_for_branch_year(student.branch, student.year)

    # Get subjects
    subjects = []
    for semester in current_semesters:
        semester_subjects = Subject.query.filter_by(
            branch=student.branch,
            semester=semester,
            is_active=True
        ).all()
        subjects.extend(semester_subjects)

    now = datetime.now()

    # Get all published tests for student's branch
    tests = Test.query.join(Subject).filter(
        Subject.branch == student.branch,
        Test.status == 'published',
        Test.is_active == True
    ).order_by(Test.available_from).all()

    # Get test attempts
    test_attempts = TestAttempt.query.filter_by(student_id=student.id).all()
    attempted_test_ids = [attempt.test_id for attempt in test_attempts]

    # Test status calculation with proper timing
    test_data = []
    for test in tests:
        attempted = test.id in attempted_test_ids

        # Check test availability based on current time
        if now >= test.available_from and now <= test.available_until:
            status = 'available'
        elif now < test.available_from:
            status = 'upcoming'
        else:
            status = 'ended'

        test_data.append({
            'test': test,
            'status': status,
            'attempted': attempted
        })

    # Calculate attendance
    subject_attendance = []
    for subject in subjects:
        total_classes = db.session.query(db.func.count(db.func.distinct(Attendance.date))).filter_by(
            subject_id=subject.id).scalar() or 0
        present_classes = Attendance.query.filter_by(
            subject_id=subject.id, student_id=student.id, status='present').count()
        percentage = round((present_classes / total_classes) * 100, 1) if total_classes > 0 else 0

        mid_term_marks = MidTermMarks.query.filter_by(
            student_id=student.id, subject_id=subject.id, exam_type='mid_term').first()

        subject_attendance.append({
            'subject': subject,
            'total_classes': total_classes,
            'present_classes': present_classes,
            'percentage': percentage,
            'status': 'Good' if percentage >= 75 else 'Average' if percentage >= 50 else 'Low',
            'mid_term_marks': mid_term_marks
        })

    # Get recent notices
    recent_notices = Notice.query.filter(
        Notice.is_active == True,
        Notice.expires_at >= datetime.now(),
        db.or_(
            Notice.target_audience == 'all',
            Notice.target_audience == 'students',
            db.and_(Notice.target_audience == 'students', Notice.branch == student.branch),
            db.and_(Notice.target_audience == 'students', Notice.branch == student.branch, Notice.year == student.year)
        )
    ).order_by(Notice.created_at.desc()).limit(5).all()

    return render_template('student/dashboard.html',
                           student=student,
                           subject_attendance=subject_attendance,
                           current_semester_config=current_semester_config,
                           current_semesters=current_semesters,
                           recent_notices=recent_notices,
                           test_data=test_data,
                           datetime=datetime)


@app.route('/student/tests')
@login_required
def student_tests():
    """Student test dashboard - FIXED TIMEZONE"""
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash('Student record not found', 'danger')
        return redirect(url_for('logout'))

    # FIXED: Use IST time for everything
    now_ist = get_ist_time()

    # Get all published tests for student's branch
    tests = Test.query.join(Subject).filter(
        Subject.branch == student.branch,
        Test.status == 'published',
        Test.is_active == True
    ).order_by(Test.available_from).all()

    test_attempts = TestAttempt.query.filter_by(student_id=student.id).all()
    attempted_test_ids = [attempt.test_id for attempt in test_attempts]

    test_data = []
    for test in tests:
        attempted = test.id in attempted_test_ids

        # FIXED: Convert test times to IST for comparison
        available_from_ist = convert_to_ist(test.available_from) if test.available_from else None
        available_until_ist = convert_to_ist(test.available_until) if test.available_until else None

        # Debug print
        print(f"TEST: {test.title}")
        print(f"Now IST: {now_ist}")
        print(f"Available From IST: {available_from_ist}")
        print(f"Available Until IST: {available_until_ist}")

        # Determine status based on IST times
        if available_from_ist and now_ist < available_from_ist:
            status = 'upcoming'
            button_text = 'Starts Soon'
            button_disabled = True
        elif available_until_ist and now_ist > available_until_ist:
            status = 'ended'
            button_text = 'Test Ended'
            button_disabled = True
        else:
            status = 'available'
            button_text = 'Start Test'
            button_disabled = False

        test_data.append({
            'test': test,
            'status': status,
            'button_text': button_text,
            'button_disabled': button_disabled,
            'attempted': attempted,
            'available_from_ist': available_from_ist,
            'available_until_ist': available_until_ist
        })

    return render_template(
        'student/tests.html',
        student=student,
        test_data=test_data,
        datetime=datetime,
        get_ist_time=get_ist_time
    )
@app.route('/debug_tests')
@login_required
def debug_tests():
    """Debug route to check test visibility"""
    if current_user.role != 'student':
        return "Admins/Professors ke liye nahi hai"

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        return "Student not found"

    now = datetime.now()

    result = f"""
    <h3>Debug Info for {student.roll}</h3>
    <p>Branch: {student.branch}</p>
    <p>Current Time: {now}</p>
    <hr>
    """

    # All tests in this branch
    tests = Test.query.join(Subject).filter(
        Subject.branch == student.branch
    ).all()

    result += f"<h4>All Tests in {student.branch} Branch: {len(tests)}</h4>"

    for test in tests:
        result += f"""
        <div style="border: 1px solid #ccc; padding: 10px; margin: 5px;">
            <strong>{test.title}</strong><br>
            Subject: {test.subject.code} - {test.subject.name}<br>
            Status: <span style="color: {'green' if test.status == 'published' else 'red'}">{test.status}</span><br>
            Active: {test.is_active}<br>
            Available: {test.available_from} to {test.available_until}<br>
            Timing: Now ({now}) is between available dates: {test.available_from <= now <= test.available_until}
        </div>
        """

    return result


@app.route('/student/test/<int:test_id>/instructions')
@login_required
def test_instructions(test_id):
    """Test instructions page - FIXED TIMEZONE"""
    if current_user.role != 'student':
        flash("Access denied!", "danger")
        return redirect(url_for('student_tests'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash("Student record not found!", "danger")
        return redirect(url_for('logout'))

    test = Test.query.get_or_404(test_id)

    # Check branch
    if test.subject.branch != student.branch:
        flash("Access to this test is denied!", "danger")
        return redirect(url_for('student_tests'))

    # FIXED: Use IST time for availability check
    now_ist = get_ist_time()
    available_from_ist = convert_to_ist(test.available_from)
    available_until_ist = convert_to_ist(test.available_until)

    print(f"INSTRUCTIONS CHECK: {test.title}")
    print(f"Now IST: {now_ist}")
    print(f"Available From IST: {available_from_ist}")
    print(f"Available Until IST: {available_until_ist}")

    # Check test availability in IST
    if now_ist < available_from_ist:
        flash("This test has not started yet!", "warning")
        return redirect(url_for('student_tests'))

    if now_ist > available_until_ist:
        flash("This test is no longer available!", "danger")
        return redirect(url_for('student_tests'))

    # Check existing attempt
    existing_attempt = TestAttempt.query.filter_by(
        test_id=test.id,
        student_id=student.id
    ).first()

    if existing_attempt and existing_attempt.submitted:
        flash("You have already attempted this test!", "info")
        return redirect(url_for('student_tests'))

    return render_template(
        "student/test_instructions.html",
        test=test,
        student=student,
        available_from_ist=available_from_ist,
        available_until_ist=available_until_ist
    )


@app.route('/student/test/<int:test_id>/start_smart', methods=['GET', 'POST'])
@login_required
def start_test_smart(test_id):
    """Start test - FIXED TIMEZONE"""
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash('Student record not found', 'danger')
        return redirect(url_for('logout'))

    test = Test.query.get_or_404(test_id)

    # FIXED: Use IST time for final check
    now_ist = get_ist_time()
    available_from_ist = convert_to_ist(test.available_from)
    available_until_ist = convert_to_ist(test.available_until)

    print(f"START TEST CHECK: {test.title}")
    print(f"Now IST: {now_ist}")
    print(f"Available From IST: {available_from_ist}")
    print(f"Available Until IST: {available_until_ist}")

    # Final availability check in IST
    if now_ist < available_from_ist:
        flash('Test has not started yet!', 'warning')
        return redirect(url_for('student_tests'))

    if now_ist > available_until_ist:
        flash('Test availability period has ended!', 'warning')
        return redirect(url_for('student_tests'))

    # Check existing attempt
    existing_attempt = TestAttempt.query.filter_by(
        student_id=student.id,
        test_id=test_id
    ).first()

    if existing_attempt:
        if existing_attempt.submitted:
            flash('You have already submitted this test', 'warning')
            return redirect(url_for('student_tests'))
        else:
            # Resume existing attempt
            attempt = existing_attempt
    else:
        # Create new attempt
        attempt = TestAttempt(
            student_id=student.id,
            test_id=test_id,
            start_time=datetime.now(),
            expected_end_time=datetime.now() + timedelta(minutes=test.duration_minutes),
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        db.session.add(attempt)
        db.session.commit()

    questions = Question.query.filter_by(test_id=test_id).order_by(Question.question_order).all()

    # Calculate remaining seconds
    now = datetime.now()
    remaining_seconds = int((attempt.expected_end_time - now).total_seconds())

    if remaining_seconds <= 0:
        if test.auto_submit:
            auto_submit_test(attempt.id)
        flash('Test time has expired', 'warning')
        return redirect(url_for('student_tests'))

    # ✅ CRITICAL FIX: Add this return statement
    return render_template('student/test_page_smart.html',
                           test=test,
                           questions=questions,
                           attempt=attempt,
                           remaining_seconds=remaining_seconds)

    # ... rest of existing start_test_smart code ...
@app.route('/student/test/<int:attempt_id>/submit_answer', methods=['POST'])
@login_required
def submit_answer(attempt_id):
    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    attempt = TestAttempt.query.get_or_404(attempt_id)
    student = Student.query.filter_by(roll=current_user.student_roll).first()

    if not student or attempt.student_id != student.id:
        return jsonify({'success': False, 'error': 'Access denied'})

    if attempt.submitted:
        return jsonify({'success': False, 'error': 'Test already submitted'})

    try:
        data = request.get_json()
        question_id = data.get('question_id')
        selected_answer = data.get('selected_answer', '').strip().upper()

        print(f" SAVE ANSWER: Attempt {attempt_id}, Q{question_id} = '{selected_answer}'")

        # [OK] FIX: Clean the answer - take only first character
        if selected_answer:
            selected_answer = selected_answer[0]  # Only take first character (A, B, C, D)

        # [OK] FIX: DELETE existing answer first (PREVENT DUPLICATES)
        deleted_count = StudentAnswer.query.filter_by(
            attempt_id=attempt_id,
            question_id=question_id
        ).delete()

        if deleted_count > 0:
            print(f" Deleted {deleted_count} existing answer(s) for Q{question_id}")

        # [OK] FIX: Create new answer (only one answer per question)
        student_answer = StudentAnswer(
            attempt_id=attempt_id,
            question_id=question_id,
            selected_answer=selected_answer
        )
        db.session.add(student_answer)
        print(f" Created new answer for Q{question_id}")

        # [OK] FIXED: Calculate marks immediately with proper comparison
        question = Question.query.get(question_id)
        if question:
            # Clean the correct answer too
            correct_answer = question.correct_answer.strip().upper() if question.correct_answer else ''

            print(f"   Correct Answer: '{correct_answer}'")
            print(f"   Selected Answer: '{selected_answer}'")

            if selected_answer and correct_answer and selected_answer == correct_answer:
                student_answer.marks_obtained = question.marks
                student_answer.is_correct = True
                print(f"[SUCCESS] CORRECT: +{question.marks} marks")
            else:
                student_answer.marks_obtained = 0
                student_answer.is_correct = False
                print(f"[ERROR] WRONG: 0 marks")
        else:
            print(f"[WARNING] Question {question_id} not found")

        # Update last activity
        attempt.last_activity = datetime.now()

        db.session.commit()
        print(f"[OK] Answer saved successfully: {selected_answer}")

        return jsonify({'success': True, 'message': 'Answer saved'})

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Error saving answer: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/student/test/<int:attempt_id>/tab_switch', methods=['POST'])
@login_required
def track_tab_switch(attempt_id):
    """Track tab switches during test"""
    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    attempt = TestAttempt.query.get_or_404(attempt_id)
    if attempt.student.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'})

    if attempt.submitted:
        return jsonify({'success': False, 'error': 'Test already submitted'})

    try:
        data = request.get_json()
        tab_switch_count = data.get('count', 0)

        attempt.tab_switch_count = tab_switch_count
        attempt.last_activity = datetime.now()

        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/student/test/<int:attempt_id>/submit_test', methods=['POST'])
@login_required
def submit_test(attempt_id):
    print(f"[SUCCESS] SUBMIT TEST CALLED: Attempt {attempt_id}")

    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    attempt = TestAttempt.query.get(attempt_id)
    if not attempt:
        return jsonify({'success': False, 'error': 'Test attempt not found'})

    # Student validation
    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student or attempt.student_id != student.id:
        return jsonify({'success': False, 'error': 'Access denied'})

    if attempt.submitted:
        return jsonify({'success': False, 'error': 'Test already submitted'})

    try:
        print("[UPDATE] SUBMIT: Starting marks calculation...")

        # [OK] FIX: Get all answers with questions
        answers = StudentAnswer.query.filter_by(attempt_id=attempt_id).all()
        print(f" SUBMIT: Found {len(answers)} answers")

        total_marks = 0
        correct_answers = 0

        # [OK] DEBUG: Pehle check karo kya data sahi hai
        print("=== DEBUG: Checking all answers ===")
        for answer in answers:
            question = Question.query.get(answer.question_id)
            if question:
                selected = answer.selected_answer.strip().upper() if answer.selected_answer else ''
                correct = question.correct_answer.strip().upper() if question.correct_answer else ''

                print(f"Q{answer.question_id}: Selected='{selected}', Correct='{correct}', Marks={question.marks}")

                # Compare answers
                if selected and correct and selected == correct:
                    marks_to_add = question.marks
                    total_marks += marks_to_add
                    correct_answers += 1
                    print(f"  [OK] CORRECT: Adding {marks_to_add} marks")
                else:
                    print(f"  [ERROR] WRONG: Adding 0 marks")
            else:
                print(f"  [WARNING] QUESTION {answer.question_id} NOT FOUND")

        print(f"[INFO] SUBMIT: Final Calculation:")
        print(f"   Total Marks: {total_marks}")
        print(f"   Correct Answers: {correct_answers}/{len(answers)}")
        print(f"   Test Total Marks: {attempt.test.total_marks}")

        # [OK] FIX: Double check - total_marks test ke total_marks se zyada na ho
        if total_marks > attempt.test.total_marks:
            print(f"[WARNING] WARNING: Calculated marks ({total_marks}) exceed test total ({attempt.test.total_marks})")
            print(f"[WARNING] This indicates duplicate answers in database")
            # Limit to test total marks and log the issue
            total_marks = min(total_marks, attempt.test.total_marks)

        # Update attempt
        attempt.end_time = datetime.now()
        attempt.total_marks_obtained = total_marks
        attempt.submitted = True

        db.session.commit()

        print("[OK] SUBMIT: Test submitted successfully!")
        print(f"[SUCCESS] Attempt {attempt_id} now has {attempt.total_marks_obtained} marks")

        return jsonify({
            'success': True,
            'message': 'Test submitted successfully!',
            'marks_obtained': total_marks,
            'total_marks': attempt.test.total_marks,
            'correct_answers': correct_answers,
            'total_questions': len(answers)
        })

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] ERROR in submit_test: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'})


@app.route('/debug/student_answers/<int:attempt_id>')
@login_required
def debug_student_answers(attempt_id):
    """Debug student answers for an attempt"""
    attempt = TestAttempt.query.get_or_404(attempt_id)

    if current_user.role != 'admin' and (current_user.role == 'student' and attempt.student.user_id != current_user.id):
        return "Access denied"

    result = f"""
    <h3>Debug Student Answers - Attempt {attempt_id}</h3>
    <p>Student: {attempt.student.name} ({attempt.student.roll})</p>
    <p>Test: {attempt.test.title}</p>
    <p>Submitted: {attempt.submitted}</p>
    <p>Marks Obtained: {attempt.total_marks_obtained}</p>
    <hr>
    <h4>All Answers:</h4>
    """

    answers = StudentAnswer.query.filter_by(attempt_id=attempt_id).all()

    if not answers:
        result += "<p class='text-danger'>No answers found!</p>"
    else:
        for answer in answers:
            question = Question.query.get(answer.question_id)
            if question:
                selected = answer.selected_answer or 'Not answered'
                correct = question.correct_answer or 'No correct answer'
                is_correct = (selected.strip().upper() == correct.strip().upper())

                result += f"""
                <div style="border: 1px solid {'#28a745' if is_correct else '#dc3545'}; padding: 10px; margin: 5px; background: {'#d4edda' if is_correct else '#f8d7da'}">
                    <strong>Q{answer.question_id}:</strong> {question.question_text[:100]}...<br>
                    <strong>Selected:</strong> <span style="color: {'green' if is_correct else 'red'}">{selected}</span> | 
                    <strong>Correct:</strong> {correct} | 
                    <strong>Marks:</strong> {answer.marks_obtained}/{question.marks} | 
                    <strong>Is Correct:</strong> {is_correct}
                </div>
                """
            else:
                result += f"<p>Question {answer.question_id} not found!</p>"

    # Calculate what total should be
    calculated_total = sum([ans.marks_obtained for ans in answers])
    result += f"""
    <hr>
    <div class="alert {'alert-success' if calculated_total == attempt.total_marks_obtained else 'alert-danger'}">
        <strong>Marks Summary:</strong><br>
        Calculated from answers: {calculated_total}<br>
        Stored in attempt: {attempt.total_marks_obtained}<br>
        Status: {'[OK] MATCH' if calculated_total == attempt.total_marks_obtained else '[ERROR] MISMATCH'}
    </div>
    """

    return result


@app.route('/debug/test-routes')
def debug_test_routes():
    """Test all test routes"""
    routes = [
        '/student/test/1/instructions',
        '/student/test/1/start_smart',
        '/student/tests',
        '/debug/routes'
    ]

    result = "<h3>Test Routes Debug</h3>"
    for route in routes:
        result += f'<p><a href="{route}" target="_blank">{route}</a></p>'

    return result


@app.route('/debug/templates')
def debug_templates():
    """Check if template files exist"""
    import os
    template_dir = 'templates/student'

    files = {
        'test_instructions.html': os.path.exists(f'{template_dir}/test_instructions.html'),
        'test_page_smart.html': os.path.exists(f'{template_dir}/test_page_smart.html')
    }

    result = "<h3>Template Files Status</h3>"
    for file, exists in files.items():
        status = "✅ EXISTS" if exists else "❌ MISSING"
        result += f"<p>{file}: {status}</p>"

    return result


@app.route('/debug/simple-test/<int:test_id>')
@login_required
def debug_simple_test(test_id):
    """Simple test page without template"""
    test = Test.query.get_or_404(test_id)
    return f"""
    <h1>Simple Test - {test.title}</h1>
    <p>This is a simple test page without template</p>
    <p>Test ID: {test_id}</p>
    <p>If this works, then template issue hai</p>
    <a href="/student/tests">Back to Tests</a>
    """
# ========== EMERGENCY ROUTES ==========

@app.route('/urgent/test')
def urgent_test():
    return "🚨 URGENT TEST - ROUTES WORKING!"

@app.route('/urgent/instructions/<int:test_id>')
def urgent_instructions(test_id):
    return f"📝 URGENT INSTRUCTIONS - Test {test_id}"

@app.route('/urgent/start/<int:test_id>')
def urgent_start(test_id):
    return f"🎯 URGENT START - Test {test_id}"
@app.route('/admin/fix_all_attempt_marks/<int:test_id>')
@login_required
def fix_all_attempt_marks(test_id):
    """Fix marks for all attempts of a test"""
    if current_user.role != 'admin':
        return "Access denied"

    attempts = TestAttempt.query.filter_by(test_id=test_id).all()
    fixed_count = 0

    result = f"<h3>Fixing marks for all attempts of test {test_id}</h3>"

    for attempt in attempts:
        answers = StudentAnswer.query.filter_by(attempt_id=attempt.id).all()
        total_marks = sum([ans.marks_obtained for ans in answers])

        if attempt.total_marks_obtained != total_marks:
            result += f"<p>Fixed attempt {attempt.id}: {attempt.total_marks_obtained} -> {total_marks}</p>"
            attempt.total_marks_obtained = total_marks
            fixed_count += 1

    if fixed_count > 0:
        db.session.commit()
        result += f"<h4 style='color: green'>[OK] Fixed {fixed_count} attempts</h4>"
    else:
        result += "<h4>No fixes needed</h4>"

    return result


@app.route('/student/notes')
@login_required
def student_notes():
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash('Student record not found', 'danger')
        return redirect(url_for('logout'))

    notes = Notes.query.join(Subject).filter(
        Subject.branch == student.branch,
        Notes.is_active == True
    ).order_by(Notes.uploaded_at.desc()).all()

    return render_template('student/notes.html', notes=notes, student=student)


@app.route('/student/notices')
@login_required
def student_notices():
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    student = Student.query.filter_by(roll=current_user.student_roll).first()
    if not student:
        flash('Student record not found', 'danger')
        return redirect(url_for('logout'))

    from datetime import datetime
    notices = Notice.query.filter(
        Notice.is_active == True,
        Notice.expires_at >= datetime.now(),
        db.or_(
            Notice.target_audience == 'all',
            Notice.target_audience == 'students',
            db.and_(Notice.target_audience == 'students', Notice.branch == student.branch),
            db.and_(Notice.target_audience == 'students', Notice.branch == student.branch, Notice.year == student.year)
        )
    ).order_by(Notice.created_at.desc()).all()

    for notice in notices:
        mark_notice_as_seen(notice.id, current_user.id)

    return render_template('student/notices.html', notices=notices, student=student)


@app.route('/download_note/<int:note_id>')
@login_required
def download_note(note_id):
    note = Notes.query.get_or_404(note_id)

    if current_user.role == 'student':
        pass
    elif current_user.role == 'professor' and note.professor_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('prof_notes'))

    if note.file_path and os.path.exists(note.file_path):
        return send_file(note.file_path, as_attachment=True, download_name=note.file_name)
    else:
        flash('File not found', 'danger')
        return redirect(url_for('student_dashboard' if current_user.role == 'student' else 'prof_notes'))


# ========== PROFILE ROUTES ==========
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    student = None
    if current_user.role == 'student':
        student = Student.query.filter_by(roll=current_user.student_roll).first()

    if request.method == 'POST':
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and file.filename != '':
                file.seek(0, os.SEEK_END)
                file_size = file.tell()
                file.seek(0)

                if file_size > MAX_FILE_SIZE:
                    flash('File size too large. Maximum 5MB allowed.', 'danger')
                elif allowed_file(file.filename):
                    if current_user.profile_photo:
                        old_file_path = os.path.join(PROFILE_PHOTOS_FOLDER, current_user.profile_photo)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)

                    filename = save_profile_photo(file, current_user.id)
                    if filename:
                        current_user.profile_photo = filename
                        db.session.commit()
                        flash('Profile photo updated successfully!', 'success')
                else:
                    flash('Invalid file type. Allowed: PNG, JPG, JPEG, GIF', 'danger')

        return redirect(url_for('profile'))

    return render_template('profile.html', student=student)


@app.route('/profile_photo/<int:user_id>')
@login_required
def get_profile_photo(user_id):
    user = User.query.get_or_404(user_id)

    if user.profile_photo and os.path.exists(os.path.join(PROFILE_PHOTOS_FOLDER, user.profile_photo)):
        return send_from_directory(PROFILE_PHOTOS_FOLDER, user.profile_photo)
    else:
        return send_from_directory('static', 'images/default-avatar.png')


# ========== TIMETABLE ROUTES ==========
@app.route('/admin/timetable')
@login_required
def admin_timetable():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('login'))

    try:
        professors = User.query.filter_by(role='professor').all()
        subjects = Subject.query.filter_by(is_active=True).all()
        allotments = ProfessorSubject.query.all()

        allotments_info = []
        for allotment in allotments:
            prof = User.query.get(allotment.professor_id)
            subj = Subject.query.get(allotment.subject_id)
            if prof and subj:
                allotments_info.append({
                    "id": allotment.id,
                    "prof": prof,
                    "subj": subj
                })

        allotments_count = len(allotments_info)

        # [OK] FIX: Empty timetables dictionary pass karo
        timetables = {}

        return render_template('admin/timetable.html',
                               professors=professors,
                               subjects=subjects,
                               allotments_info=allotments_info,
                               allotments_count=allotments_count,
                               timetables=timetables)  # [OK] Yeh line add karo

    except Exception as e:
        print(f"[ERROR] Error in admin_timetable: {e}")
        flash('Error loading timetable page', 'danger')
        return redirect(url_for('admin_dashboard'))


@app.route('/generate_timetable', methods=['POST'])
@login_required
def generate_timetable():
    if current_user.role != 'admin':
        flash("Access denied!", "danger")
        return redirect(url_for('login'))

    try:
        # Get form selections
        branches = request.form.getlist("branches")
        years = [int(y) for y in request.form.getlist("years")]
        semesters = [int(s) for s in request.form.getlist("semesters")]

        if not branches or not years or not semesters:
            flash("Please select at least one branch, year and semester", "warning")
            return redirect(url_for('admin_timetable'))

        print(f"[START] Timetable requested => Branches: {branches}, Years: {years}, Semesters: {semesters}")

        # Fetch only professors
        available_faculties = User.query.filter_by(role="professor").all()

        if not available_faculties:
            flash("No professors found. Please add professors first.", "danger")
            return redirect(url_for('admin_timetable'))

        print(f"[INFO] Professors available => {[f.id for f in available_faculties]}")

        # Ensure subjects allotted
        faculty_ids = [f.id for f in available_faculties]
        assigned_subjects = ProfessorSubject.query.filter(
            ProfessorSubject.professor_id.in_(faculty_ids)
        ).all()

        if not assigned_subjects:
            flash("No subject allotted to any professor! Please allot subjects first.", "warning")
            return redirect(url_for('admin_timetable'))

        print(f"[INFO] Subject allotments found => {len(assigned_subjects)} records")

        # Generate timetable
        timetables = generate_smart_timetable(branches, years, semesters)

        if timetables:
            print(f"[INFO] Timetables created => Count: {len(timetables)}")

            result = save_timetable_to_db(timetables)

            if result:
                flash("Timetable generated successfully!", "success")
                return redirect(url_for('view_combined_timetable',
                                        branches=",".join(branches),
                                        years=",".join(map(str, years)),
                                        semesters=",".join(map(str, semesters))))
            else:
                flash("Database save failed!", "danger")
        else:
            flash("Timetable generation failed — Check subject allotment & professor availability.", "danger")

    except Exception as e:
        print(f"[ERROR] Timetable generation crash: {e}")
        import traceback
        traceback.print_exc()

        flash(f"Internal error: {str(e)}", "danger")

    return redirect(url_for('admin_timetable'))

@app.route('/timetable/combined')
@login_required
def view_combined_timetable():
    branches = request.args.get('branches', 'CSE,AD').split(',')
    years = [int(y) for y in request.args.get('years', '3').split(',')]
    semesters = [int(s) for s in request.args.get('semesters', '5,6').split(',')]

    timetables = {}
    for branch in branches:
        for year in years:
            for semester in semesters:
             key = f"{branch}_{year}_{semester}"
             timetables[key] = get_timetable_from_db(branch, year, semester)

    return render_template('timetable/combined.html',
                           timetables=timetables,
                           years=years,
                           semesters=semesters,
                           branches=branches)

# ========== AUTO-SUBMIT BACKGROUND TASK ==========
@app.route('/cron/auto_submit_tests')
def auto_submit_expired_tests():
    """Background task to auto-submit expired tests"""
    try:
        now = datetime.now()

        expired_attempts = TestAttempt.query.filter(
            TestAttempt.submitted == False,
            TestAttempt.expected_end_time <= now
        ).all()

        submitted_count = 0
        for attempt in expired_attempts:
            auto_submit_test(attempt.id)
            submitted_count += 1

        return jsonify({'success': True, 'submitted_count': submitted_count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/student/test/<int:attempt_id>/check_time', methods=['POST'])
@login_required
def check_test_time(attempt_id):
    """Check if test time has expired"""
    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    attempt = TestAttempt.query.get_or_404(attempt_id)
    if attempt.student.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'})

    if attempt.submitted:
        return jsonify({'success': False, 'expired': True, 'message': 'Test already submitted'})

    now = datetime.now()

    if now > attempt.expected_end_time:
        # Auto submit the test
        if attempt.test.auto_submit:
            auto_submit_test(attempt.id)
            return jsonify({'success': True, 'expired': True, 'message': 'Test auto-submitted due to time expiry'})
        else:
            return jsonify({'success': True, 'expired': True, 'message': 'Test time expired'})

    remaining_seconds = int((attempt.expected_end_time - now).total_seconds())
    return jsonify({
        'success': True,
        'expired': False,
        'remaining_seconds': remaining_seconds
    })


@app.route('/student/test/<int:attempt_id>/auto_save', methods=['POST'])
@login_required
def auto_save_answer(attempt_id):
    """Auto-save answer without validation"""
    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    attempt = TestAttempt.query.get_or_404(attempt_id)
    if attempt.student.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'})

    if attempt.submitted:
        return jsonify({'success': False, 'error': 'Test already submitted'})

    try:
        data = request.get_json()
        question_id = data.get('question_id')
        selected_answer = data.get('selected_answer', '').strip().upper()

        student_answer = StudentAnswer.query.filter_by(
            attempt_id=attempt_id,
            question_id=question_id
        ).first()

        if student_answer:
            student_answer.selected_answer = selected_answer
        else:
            student_answer = StudentAnswer(
                attempt_id=attempt_id,
                question_id=question_id,
                selected_answer=selected_answer
            )
            db.session.add(student_answer)

        # Update last activity
        attempt.last_activity = datetime.now()

        db.session.commit()
        return jsonify({'success': True, 'message': 'Answer auto-saved'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/reset_migration')
def reset_migration():
    """Reset migration version manually"""
    try:
        with db.engine.connect() as conn:
            # Check if alembic_version table exists
            result = conn.execute(text("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='alembic_version'
            """))

            if result.fetchone():
                # Table exists - delete the version
                conn.execute(text("DELETE FROM alembic_version"))
                conn.commit()
                return "[OK] Migration version reset successfully! Now run: flask db migrate"
            else:
                # Table doesn't exist - create it
                conn.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
                conn.commit()
                return "[OK] Alembic version table created! Now run: flask db migrate"

    except Exception as e:
        return f"[ERROR] Error: {str(e)}"


@app.route('/student/test/<int:test_id>/verify_security_code', methods=['POST'])
@login_required
def verify_security_code(test_id):
    if current_user.role != 'student':
        return jsonify({'success': False, 'error': 'Access denied'})

    test = Test.query.get_or_404(test_id)

    try:
        data = request.get_json()
        entered_code = data.get('security_code', '').strip()

        print(f"[SUCCESS] Security Code Verification: Test {test_id}, Code: {entered_code}")

        # Check if security code is required
        if not test.require_security_code:
            return jsonify({'success': True, 'message': 'Security not required'})

        # Verify code
        if test.security_code == entered_code:
            # Code verified - mark as verified
            test.security_code_verified = True
            db.session.commit()
            print(f"[OK] Security code verified for test {test_id}")
            return jsonify({'success': True, 'message': 'Code verified'})
        else:
            print(f"[ERROR] Invalid security code for test {test_id}")
            return jsonify({'success': False, 'error': 'Invalid security code'})

    except Exception as e:
        print(f"[ERROR] Error in security verification: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/admin/migrate_data')
@login_required
def migrate_data():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        # Get all data from SQLite
        from sqlalchemy import create_engine
        import pandas as pd

        # SQLite connection
        sqlite_engine = create_engine('sqlite:///college_attendance.db')

        # PostgreSQL connection (from environment variable)
        postgres_url = os.environ.get('DATABASE_URL')
        if postgres_url.startswith('postgres://'):
            postgres_url = postgres_url.replace('postgres://', 'postgresql+psycopg://', 1)
        postgres_engine = create_engine(postgres_url)

        # List of tables to migrate
        tables = ['user', 'student', 'subject', 'professor_subject', 'attendance',
                  'attendance_report', 'mid_term_marks', 'notes', 'notice', 'test',
                  'question', 'test_attempt', 'student_answer']

        migrated_tables = []

        for table in tables:
            try:
                # Read data from SQLite
                df = pd.read_sql_table(table, sqlite_engine)

                # Write to PostgreSQL
                df.to_sql(table, postgres_engine, if_exists='replace', index=False)

                migrated_tables.append(f"✅ {table}: {len(df)} records")

            except Exception as e:
                migrated_tables.append(f"❌ {table}: {str(e)}")

        # Update database configuration to use PostgreSQL
        app.config['SQLALCHEMY_DATABASE_URI'] = postgres_url

        result = "<br>".join(migrated_tables)
        flash('Data migration completed!', 'success')
        return f"""
        <h3>Data Migration Results</h3>
        {result}
        <br><br>
        <a href="/admin" class="btn btn-primary">Back to Dashboard</a>
        """

    except Exception as e:
        flash(f'Migration failed: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))


import json
from sqlalchemy import text


@app.route('/admin/import_data')
@login_required
def import_data():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        # Load exported data
        with open('database_export.json', 'r', encoding='utf-8') as f:
            export_data = json.load(f)

        results = []

        # Import data table by table
        for table_name, records in export_data.items():
            try:
                if table_name == 'user':
                    for record in records:
                        # Check if user already exists
                        existing = User.query.filter_by(email=record['email']).first()
                        if not existing:
                            user = User(
                                username=record['username'],
                                email=record['email'],
                                fullname=record['fullname'],
                                role=record['role'],
                                branch=record['branch'],
                                student_roll=record.get('student_roll'),
                                email_verified=record.get('email_verified', True),
                                is_active=record.get('is_active', True)
                            )
                            # Note: Password cannot be imported (security)
                            user.set_password('temp123')  # Set temporary password
                            db.session.add(user)
                    db.session.commit()
                    results.append(f"✅ Users: {len(records)} imported")

                elif table_name == 'student':
                    for record in records:
                        existing = Student.query.filter_by(roll=record['roll']).first()
                        if not existing:
                            student = Student(
                                roll=record['roll'],
                                name=record['name'],
                                branch=record['branch'],
                                year=record['year']
                            )
                            db.session.add(student)
                    db.session.commit()
                    results.append(f"✅ Students: {len(records)} imported")

                elif table_name == 'subject':
                    for record in records:
                        existing = Subject.query.filter_by(code=record['code']).first()
                        if not existing:
                            subject = Subject(
                                code=record['code'],
                                name=record['name'],
                                branch=record['branch'],
                                semester=record['semester'],
                                is_active=record.get('is_active', True)
                            )
                            db.session.add(subject)
                    db.session.commit()
                    results.append(f"✅ Subjects: {len(records)} imported")

                elif table_name == 'professor_subject':
                    for record in records:
                        # Check if both professor and subject exist
                        professor = User.query.get(record['professor_id'])
                        subject = Subject.query.get(record['subject_id'])

                        if professor and subject:
                            existing = ProfessorSubject.query.filter_by(
                                professor_id=record['professor_id'],
                                subject_id=record['subject_id']
                            ).first()

                            if not existing:
                                allotment = ProfessorSubject(
                                    professor_id=record['professor_id'],
                                    subject_id=record['subject_id']
                                )
                                db.session.add(allotment)
                    db.session.commit()
                    results.append(f"✅ Professor Subjects: {len(records)} imported")

                # Add more tables as needed...

            except Exception as e:
                results.append(f"❌ {table_name}: Error - {str(e)}")
                db.session.rollback()

        # Create admin notice
        notice = Notice(
            title="Data Import Completed",
            message=f"Data import completed on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            created_by=current_user.id,
            target_audience="all",
            is_important=True
        )
        db.session.add(notice)
        db.session.commit()

        result_html = "<br>".join(results)
        return f"""
        <h3>Data Import Results</h3>
        {result_html}
        <br><br>
        <a href="/admin" class="btn btn-primary">Back to Dashboard</a>
        """

    except Exception as e:
        flash(f'Import failed: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))


@app.route('/test_brevo_transactional')
def test_brevo_transactional():
    """Test Brevo Transactional Email API"""
    test_email = "ravikumarmohane@gmail.com"
    test_otp = generate_otp()

    print("🧪 Testing Brevo Transactional API...")
    success = send_otp_via_brevo_api(test_email, test_otp)

    return f"""
    <h3>Brevo Transactional API Test</h3>
    <p><strong>Email:</strong> {test_email}</p>
    <p><strong>OTP:</strong> {test_otp}</p>
    <p><strong>Result:</strong> {'✅ SUCCESS' if success else '❌ FAILED'}</p>
    <p><strong>Check:</strong></p>
    <ol>
        <li>Server console for detailed logs</li>
        <li>Brevo Dashboard → Transactional Emails</li>
        <li>Your email inbox + spam folder</li>
    </ol>
    """

# ========== MAIN APPLICATION LAUNCH ==========
if __name__ == '__main__':
    # Get port from environment variable or default to 5000
    port = int(os.environ.get("PORT", 5000))

    # Run the app
    app.run(
        host='0.0.0.0',  # Important: Bind to all interfaces
        port=port,
        debug=False  # Set to False in production
    )